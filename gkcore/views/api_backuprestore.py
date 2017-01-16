"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@gmail.com>
"Prajkta Patkar" <prajkta.patkar007@gmail.com>

"""

from gkcore import eng, enumdict
from gkcore.views.api_login import authCheck
from gkcore.models.gkdb import organisation,accounts,users,bankrecon,categorysubcategories,categoryspecs,customerandsupplier,dcinv,delchal,godown,goprod,groupsubgroups,invoice,projects,product,purchaseorder,transfernote,stock,tax,unitofmeasurement,vouchers,voucherbin
from sqlalchemy.sql import select
from datetime import datetime,date
from openpyxl import Workbook
from openpyxl.styles import Font
import json
from sqlalchemy.engine.base import Connection
from sqlalchemy import and_, exc
from pyramid.request import Request
from pyramid.response import Response
from pyramid.view import view_defaults,  view_config
from sqlalchemy.ext.baked import Result
import gkcore
import os
from sqlalchemy.sql.functions import func
import base64
from datetime import datetime
import user

@view_defaults(route_name='backuprestore')
class api_backuprestore(object):
	def __init__(self,request):
		self.request = Request
		self.request = request
		self.con = Connection
		print "backup initialized"
		
					
	@view_config(request_method='GET',renderer='json',request_param='fulldb=0')
	def backuporg(self):
		""" Purpose:
		This function gives a backup of all data for an entire organisation for a given financial year.
		Returns a spreadsheet in encoded form with list of accounts under their groups or subgroups in the first sheet.
		2nd sheet has all projects details , 3rd sheet with all vouchers and 4th sheet contains all users details.
		description:
		this function will take orgcode and userid from the get request.
		Once this data is taken from the json_body the processing is done for that organisation.
		Firstly check for the user's role.  If it is admin then,
		Query to organisation table and get organisation name , yearstart and yearend ; Fill this data into first row. 
		First it gets the list of all main groups.
		Then a main loop runs through this list.
		For each iteration a cell is created with the current groupname in bold.
		after the cell is created,
		In this loop we first check if any accounts exist for the main group.
		if row count is > 0 for the select query,
		then a loop is run to add every account name one below the other in successive cells.
		The account name is in italics.
		now query for list of subgroup for this group. 
		then if there are subgroups for this group using groupcode in the select query,
		run a for loop.
		This means if the rowcount is more than 0 then a for loop is run for these subgroups.
		At the beginning of this subgroup loop the name of the subgroup is put exactly under the group if no accounts were directly in the group,
		Else the subgroup will be put under the cell containing the last account for the main group.
		Inside this for loop we check for the list of accounts.
		If there are accounts (check for row count after query ) then run another loop inside.
		All that this loop will do is to add account name in cells below the subgroup.
		Note that account names will be in italics.
		
		Once all accounts entries have been added to first sheet next sheet will be created for project existed for organisation.
		query the projects table to get all project's name and sanctioned amount.
		
		Then next sheet create for vouchers entry.
		Now query for vouchers details such as voucherdate , vouchernumber , vouchertype , narration ,crs and drs , projectcode.
		if rowcount > 0 , fetch all data which query has return.
		loop through data & Add data in the rows and columns accordingly.
		Format decided for adding data to the spreadsheet is as follows VoucherNumber ,Date, VoucherType , DebitAccount , Debit amount,CreditAccount , Credit amount ,Narration , Lockflag , Projectname; 
		Title for each field is mentioned in 1st row.
		Filling details(values) start from column 1 and fill up the columns upto column 10 of 2nd row respectively. Row counter will increase by 1 as soon as all details in current in row is filled up.
		To get accountname for creditaccount and debitaccount firstly get accountcode form dictionary crs and drs .
		query to accounts table using accountcode and retrieve accountname.
		For multiple crs and drs for each entry , Entries for accountname and amount will be added to immediate next row .
		Projectname can be find using projectcode. 
		If projects exist project column will be created and find the project name for that voucher entry else field will be blank.
		
		Atlast 4th sheet is created if users exist other than admin, and users details are filled in the columns accordingly.
		At last Save the file in xlsx format by giving suitable name.
		encode the file using base64 encode format , now file has been converted into encoded string format, So that we can use it as value to JSON dictionary.
				
		"""
		try:
			token = self.request.headers["gktoken"]
		except:
			return  {"gkstatus":  gkcore.enumdict["UnauthorisedAccess"]}
		authDetails = authCheck(token)
		if authDetails["auth"] == False:
			return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
		else:
#			try:
				self.con = eng.connect()
				user=self.con.execute(select([users.c.userrole]).where(users.c.userid == authDetails["userid"] ))
				userRole = user.fetchone()
				if userRole[0]==-1:
					#create a workbook
					#open the book with one sheet.
					#then we will get list of all groups.
					gkwb = Workbook()
					accountList = gkwb.active
					accountList.title = "Account List"
					accountList.column_dimensions["A"].width = 80
					accountList.column_dimensions["B"].width = 30
					
					orgInfo = self.con.execute(select([organisation.c.orgname,organisation.c.orgtype,organisation.c.yearstart,organisation.c.yearend]).where(organisation.c.orgcode == authDetails["orgcode"] ))
					org = orgInfo.fetchone()
					accountList.cell(row=1,column=1,value= org["orgname"])
					accountList.cell(row=1,column=2,value= org["orgtype"])
					accountList.cell(row=1,column=3,value= str(org["yearstart"].strftime('%d-%m-%Y')))
					accountList.cell(row=1,column=4,value= str(org["yearend"].strftime('%d-%m-%Y')))
													
					mainGroups = self.con.execute(select([groupsubgroups.c.groupcode, groupsubgroups.c.groupname]).where(and_(groupsubgroups.c.orgcode == authDetails["orgcode"], groupsubgroups.c.subgroupof == None )))
					groups =  mainGroups.fetchall()
					cellCounter = 3
					for group in groups:
						#create first row with cell containing groupname.
						#make it bold style with font object and then go for it's accounts first.
						c = accountList.cell(row=cellCounter,column=1,value=group["groupname"])
						c.font = Font(name=c.font.name,bold=True)
						cellCounter = cellCounter + 1
						grpaccounts = self.con.execute(select([accounts.c.accountname,accounts.c.openingbal]).where(and_(accounts.c.groupcode == group["groupcode"],accounts.c.orgcode == authDetails["orgcode"]) ))
						if grpaccounts.rowcount > 0:
							account = grpaccounts.fetchall()
							for acct in account:
								a = accountList.cell(row=cellCounter,column=1,value= acct["accountname"])
								a.font = Font(name=a.font.name,italic=True) 
								ob = accountList.cell(row=cellCounter,column=2,value= "%.2f"%float(acct["openingbal"]))
								cellCounter = cellCounter + 1
						#search for subgroups existing for main group , create row with cell containing subgroup
						#then search for accounts existing for subgroup. and create new cell immediately under subgroup.
						subgrp = self.con.execute(select([groupsubgroups.c.groupcode, groupsubgroups.c.groupname]).where(and_(groupsubgroups.c.orgcode == authDetails["orgcode"], groupsubgroups.c.subgroupof ==group["groupcode"])))
						if subgrp.rowcount > 0:
							subgroup = subgrp.fetchall()
							for sg in subgroup:
								s = accountList.cell(row=cellCounter,column=1,value=sg["groupname"])
								cellCounter = cellCounter + 1
								grpaccounts = self.con.execute(select([accounts.c.accountname,accounts.c.openingbal]).where(and_(accounts.c.groupcode == sg["groupcode"],accounts.c.orgcode == authDetails["orgcode"]) ))
								if grpaccounts.rowcount > 0:
									account = grpaccounts.fetchall()
									for acct in account:
										a = accountList.cell(row=cellCounter,column=1,value= acct["accountname"])
										a.font = Font(name=a.font.name,italic=True) 
										ob = accountList.cell(row=cellCounter,column=2,value="%.2f"%float( acct["openingbal"]))
										cellCounter = cellCounter + 1
										
					# Now create project sheet to backup project data if any project exists in the GNUKhata.
					projInfo = self.con.execute(select([projects.c.projectname,projects.c.sanctionedamount]).where(projects.c.orgcode == authDetails["orgcode"] ))
					if projInfo.rowcount > 0:
						Projects = gkwb.create_sheet(title="Projects")
						Projects.column_dimensions["A"].width = 60
						Projects.column_dimensions["B"].width = 40
					   	Projects.cell(row= 1,column=1,value= "ProjectName")
						Projects.cell(row= 1,column=2,value= "SanctionedAmount")
					
						projrow = 2
						projData = projInfo.fetchall()
						for prj in projData :
							Projects.cell(row= projrow ,column=1,value= prj["projectname"]) 
							Projects.cell(row= projrow ,column=2,value= "%.2f"%float(prj["sanctionedamount"]))
							projrow = projrow + 1
																				
					
					voucher = self.con.execute(select([vouchers.c.vouchernumber,vouchers.c.voucherdate,vouchers.c.narration,vouchers.c.vouchertype,vouchers.c.drs,vouchers.c.crs,vouchers.c.lockflag,vouchers.c.projectcode,vouchers.c.delflag]).where(vouchers.c.orgcode== authDetails["orgcode"]))
					if voucher.rowcount > 0:
						Vouchers = gkwb.create_sheet(title="Vouchers")
						Vouchers.column_dimensions["D"].width = 25 
						Vouchers.column_dimensions["F"].width = 25
						Vouchers.column_dimensions["H"].width = 20
						
						Vouchers.cell(row= 1,column=1,value= "VchNo")
						Vouchers.cell(row= 1,column=2,value= "Date")
						Vouchers.cell(row= 1,column=3,value= "VchType")
						Vouchers.cell(row= 1,column=4,value= "DrAcc")
						Vouchers.cell(row= 1,column=5,value= "DrAmt")
						Vouchers.cell(row= 1,column=6,value= "CrAcc")
						Vouchers.cell(row= 1,column=7,value= "CrAmt")
						Vouchers.cell(row= 1,column=8,value= "Narration")
						Vouchers.cell(row= 1,column=9,value= "Lockflag")
												
						rowcounter = 2	
						if voucher.rowcount > 0:
							voucherData = voucher.fetchall()
							for vch in voucherData:
								Vouchers.cell(row= rowcounter,column=1,value=vch["vouchernumber"]) 
								Vouchers.cell(row= rowcounter,column=2,value= str(vch["voucherdate"].date().strftime('%d-%m-%Y'))) 
								Vouchers.cell(row= rowcounter,column=3,value=vch["vouchertype"])
								dr = vch["drs"]
								drcounter = rowcounter
								for draccno in dr.keys():
									draccname = self.con.execute(select([accounts.c.accountname]).where(and_(accounts.c.accountcode == draccno , accounts.c.orgcode == authDetails["orgcode"])))
									dracctname = draccname.fetchone()
									drAccName =  Vouchers.cell(row= drcounter,column=4,value=dracctname["accountname"])
									drValue = Vouchers.cell(row= drcounter,column=5,value= "%.2f"%float(dr[draccno]))
									drcounter = drcounter + 1
									
								cr = vch["crs"]
								crcounter = rowcounter
								for craccno in cr.keys():
									craccname = self.con.execute(select([accounts.c.accountname]).where(and_(accounts.c.accountcode == craccno,accounts.c.orgcode == authDetails["orgcode"])))
									cracctname = craccname.fetchone()
									crAccName = Vouchers.cell(row= crcounter,column=6,value=cracctname["accountname"])
									crvalue = Vouchers.cell(row= crcounter,column=7,value= "%.2f"%float(cr[craccno]))
									crcounter = crcounter + 1
								
								Vouchers.cell(row= rowcounter,column=8,value=vch["narration"])
								Vouchers.cell(row= rowcounter,column=9,value=vch["lockflag"])
																
								if vch["projectcode"] !=None:
									Vouchers.column_dimensions["J"].width = 20
									Vouchers.cell(row= 1,column=10,value= "Project")
									prj = self.con.execute(select([projects.c.projectname]).where(and_(projects.c.projectcode == vch["projectcode"], projects.c.orgcode == authDetails["orgcode"])))
									prjname = prj.fetchone()
									Vouchers.cell(row= rowcounter,column=10,value=prjname["projectname"])
																			
								if drcounter >= rowcounter:
									rowcounter = drcounter + 1
								else :
									rowcounter = crcounter + 1
				
					userInfo = self.con.execute(select([users.c.username,users.c.userpassword,users.c.userrole,users.c.userquestion,users.c.useranswer,]).where(and_(users.c.orgcode == authDetails["orgcode"],users.c.userrole != -1) ))
					if userInfo.rowcount > 0:
						User = gkwb.create_sheet(title="Users")
						User.column_dimensions["A"].width =30
						User.cell(row= 1,column=1,value= "UserName")
						User.cell(row= 1,column=2,value= "Password")
						User.cell(row= 1,column=3,value= "Role")
						User.cell(row= 1,column=4,value= "Question")
						User.cell(row= 1,column=5,value= "Answer")
						usRow = 2
						userData = userInfo.fetchall()
						for us in userData:
							User.cell(row = usRow ,column=1,value= us[0])
							User.cell(row = usRow ,column=2,value= us[1])
							if us[2] == 0:
								userRole = "Manager"
							else:
								userRole = "Operator" 
							User.cell(row = usRow ,column=3,value= userRole)	
							User.cell(row = usRow ,column=4,value= us[3])
							User.cell(row = usRow ,column=5,value= us[4])
							usRow = usRow + 1
					gkwb.save(filename = "/tmp/GkExport.xlsx")
					gkarch = open("/tmp/GkExport.xlsx","r")
					archData = base64.b64encode(gkarch.read())
					gkarch.close()
					return {"gkstatus":enumdict["Success"],"gkdata":archData}
#			except:
#				return {"gkstatus":gkcore.enumdict["ConnectionFailed"]}
#			finally:
#				self.con.close()
				 
				
				
	
	
						

					
			
		
		
		  

			
					
