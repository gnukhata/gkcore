# from pyramid.view import view_config, view_defaults
import io
import logging

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from pyramid.response import Response
from sqlalchemy.engine import Connection
from sqlalchemy.sql import select

from gkcore import eng
from gkcore.models.gkdb import groupsubgroups
from gkcore.models.gkdb import accounts as accounts_table
from gkcore.models.gkdb import organisation as org_table
from gkcore.models.meta import gk_api
from gkcore.views.api_gkuser import authCheck, getUserRole


def export_ledger(self):
    """Export organisation's accounts & vouchers

    `params:`
    *yearstart* = mm-dd-yyyy
    yearend= yyyy-mm-dd
    """
    self.con = Connection
    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        gkwb = Workbook()
        accountList = gkwb.active
        accountList.title = "Account List"
        accountList.column_dimensions["A"].width = 80
        accountList.column_dimensions["B"].width = 30
        gsResult = gk_api("/groupsubgroups", header, self.request)
        groupList = gsResult["gkresult"]
        ob = accountList.cell(row=1, column=2, value="Opening Balance")
        cellCounter = 2
        for group in groupList:
            g = accountList.cell(row=cellCounter, column=1, value=group["groupname"])
            g.font = Font(name=g.font.name, bold=True)
            cellCounter = cellCounter + 1
            accResult = gk_api(
                "/accounts?accbygrp&groupcode=%s" % group["groupcode"],
                header,
                self.request,
            )
            accList = accResult["gkresult"]
            for acc in accList:
                a = accountList.cell(
                    row=cellCounter, column=1, value=(acc["accountname"])
                )
                a.font = Font(name=g.font.name, italic=True)
                ob = accountList.cell(
                    row=cellCounter,
                    column=2,
                    value=float("%.2f" % float(acc["openingbal"])),
                )
                accountList.cell(row=cellCounter, column=2).number_format = "0.00"
                cellCounter = cellCounter + 1
            sgResult = gk_api(
                "/groupDetails/%s" % (group["groupcode"]), header, self.request
            )
            subgrpList = sgResult["gkresult"]
            for sg in subgrpList:
                sbg = accountList.cell(
                    row=cellCounter, column=1, value=sg["subgroupname"]
                )
                cellCounter = cellCounter + 1
                accsgResult = gk_api(
                    "/accounts?accbygrp&groupcode=%s" % sg["groupcode"],
                    header,
                    self.request,
                )
                accListsg = accsgResult["gkresult"]
                for accsg in accListsg:
                    a = accountList.cell(
                        row=cellCounter, column=1, value=(accsg["accountname"])
                    )
                    a.font = Font(name=g.font.name, italic=True)
                    ob = accountList.cell(
                        row=cellCounter,
                        column=2,
                        value=float("%.2f" % float(accsg["openingbal"])),
                    )
                    accountList.cell(row=cellCounter, column=2).number_format = "0.00"

                    cellCounter = cellCounter + 1

        Voucher = gkwb.create_sheet()
        Voucher.title = "List of all vouchers"
        yearStart = str(self.request.params["yearstart"])
        yearEnd = str(self.request.params["yearend"])
        vchResult = gk_api(
            "/transaction?getdataby=date&from=%s&to=%s" % (yearStart, yearEnd),
            header,
            self.request,
        )
        vchList = vchResult["gkresult"]
        rowCounter = crRowCounter = counter = mCounter = 1

        #  1= Date, 2 = Particulars 5 = vchtype 6 = vchno 7 = debit 8 = credit.(column no. = Column title).
        # set size of column.
        Voucher.column_dimensions["A"].width = 15
        Voucher.column_dimensions["B"].width = 40
        Voucher.column_dimensions["E"].width = 10
        Voucher.column_dimensions["F"].width = 10
        Voucher.column_dimensions["G"].width = 10
        Voucher.column_dimensions["H"].width = 10
        # Columns title
        Voucher.cell(row=rowCounter, column=1, value="Date")
        Voucher.cell(row=rowCounter, column=2, value="Particulars")
        Voucher.cell(row=rowCounter, column=5, value="Type")
        Voucher.cell(row=rowCounter, column=6, value="Voucher No.")
        Voucher.cell(row=rowCounter, column=7, value="Debit")
        Voucher.cell(row=rowCounter, column=8, value="Credit")
        # increase row counter so that data writing will start from next row
        rowCounter = rowCounter + 1
        # loop through all data(vouchers)
        for v in vchList:
            # write data in their resepective cell, & which belong to same row.
            Voucher.cell(row=rowCounter, column=1, value=v["voucherdate"])
            Voucher.cell(row=rowCounter, column=5, value=v["vouchertype"])
            Voucher.cell(row=rowCounter, column=6, value=v["voucherno"])
            Crs = v["crs"]
            Drs = v["drs"]
            # loop through drs & crs dictionaries, there can be multiple crs / drs prsent . write account name in 2nd column & debit , credit in column 7 & 8 respectively.
            for k in Drs:
                Voucher.cell(row=rowCounter, column=2, value=k)
                Voucher.cell(row=rowCounter, column=7, value="%.2f" % float(Drs[k]))
                rowCounter = rowCounter + 1
            for k in Crs:
                Voucher.cell(row=rowCounter, column=2, value=k)
                Voucher.cell(row=rowCounter, column=8, value="%.2f" % float(Crs[k]))
                rowCounter = rowCounter + 1
                # write narration after all particular details is wrote , font style & size is set different for better presentation
            if v["narration"] != "":
                a = Voucher.cell(row=rowCounter, column=2, value=v["narration"])
                a.font = Font(italic=True, size=10)
                rowCounter = rowCounter + 1
        output = io.BytesIO()
        gkwb.save(output)
        contents = output.read()
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=AllLedger.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=AllLedger.xlsx','Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print(e)
        return {"gkstatus": 3}


def import_tally(self):
    """
     This function will take a spreadsheet containing data from tally or GNUKhata
     Then the code will read the file using parsing library (openpyxl).
     With a number of post calls to REST API, the data is added to GNUKhata.
     The data consists of :
     *new subgroups if they don't exist,
     *new accounts under existing or new subgroups
     * new accounts undr group as per data provided.
     The data   should be in the following format.
     * first sheet must contain the list of accounts
     * Structure should be groups with their optional subgroups
     * if accounts are to be under a group then they should come immediately below the group
     * if there are subgroups under the group they should imediately follow the group
     * groups are in bold
     *accounts are italics
     * subgroups are normal
     * list of groups should be exactly as per GNUKhata (13 at the most ).
     The the code will move to the next sheet which contains all the vouchers.
     This time the function will take list of rows.

     Following is the structure of voucher presentation.
     Column 1 = Date , Column 2 = Contains accounts / narration when Drs & Crs do not have value , Column 5 = Voucher Type, Column 6 = Voucher no (to be ignored) ,Column 7 = Dr Amount , Column 8 = Dr amount. In this pattern all the vouchers will be taken & send to the core engine for data entry by making successive POST request.

     Take all rows from sheet. Loop through all rows till you get 'Date' word in the first column, Then set the current row index to the next row and exit the loop.
    Start a new loop for list of vouchers and start checking following conditions :
     *If 1st, 2nd & 3rd columns don't have data then continue.
     * If 1st column & 5th column contains data then pick up the date and voucher type respectively also pick up account name and debit amount from 2nd & 6th columns respectively.
     * if 1st and (5th or 6th) column contains data then it is another debit  or credit entry. In this very condition have 2 nested conditions.
     * If the 5th coulmn has data then data from the 2nd & 5th column gets added to DR dictionary.
     * If the 6th coulmn has data then data from the 2nd & 6th column gets added to CR dictionary.
     * Finally if 2nd column has data but 5th and 6th column does not pickup data from 2nd column as narration.

    """
    # First we will get list of existing groups and subgroups for this organisation.
    # we will of course lead the workbook from the request.
    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        xlsxfile = self.request.POST["gkfile"].file
        wbTally = load_workbook(xlsxfile)
        wbTally._active_sheet_index = 0
        accountSheet = wbTally.active
        accountList = tuple(accountSheet.rows)
        gsResult = gk_api("/groupsubgroups?groupflatlist", header, self.request)
        groups = gsResult["gkresult"]
        curgrpid = None
        parentgroupid = None
        parentgroup = ""
        openingBl = 0.00
        self.con = eng.connect()

        # gather user info
        user = authCheck(header["gktoken"])
        user_role = getUserRole(user["userid"], user["orgcode"])["gkresult"]["userrole"]

        # only admin can import data
        if user_role != -1:
            return {"gkstatus": 4}

        for accRow in accountList:
            if accRow[0].value == None:
                continue
            if accRow[0].font.b:
                curgrpid = groups[accRow[0].value.strip()]
                parentgroupid = groups[accRow[0].value.strip()]
                parentgroup = accRow[0].value.strip()
                continue
            if accRow[0].font.b == False and accRow[0].font.i == False:
                if accRow[0].value in groups:
                    curgrpid = groups[accRow[0].value.strip()]
                else:
                    try:
                        newsub = eng.connect().execute(
                            groupsubgroups.insert(),
                            {
                                "groupname": accRow[0].value,
                                "subgroupof": parentgroupid,
                                "orgcode": user["orgcode"],
                            },
                        )
                        # TODO check
                        curgrpid = newsub["gkresult"]
                    except Exception as e:
                        print("exception: ", e)
                    print(1, curgrpid)
            if accRow[0].font.i:
                if len(accRow) > 2:
                    if accRow[1].value == None and accRow[2].value == None:
                        try:
                            newacc = eng.connect.execute(
                                accounts_table.insert(),
                                {
                                    "accountname": accRow[0].value,
                                    "groupcode": curgrpid,
                                    "openingbal": 0.00,
                                },
                            )

                        except Exception as e:
                            print("exception: ", e)
                        continue
                    # checking if opening Balance is not in Debit column. i.e. column no. 2 (B).
                    # It means value is in credit column
                    if accRow[1].value == None and accRow[2].value != None:
                        openingBl = float(accRow[2].value)
                        # Check parent group so that opening balance type (cr/dr) can be determined.
                        if (
                            parentgroup == "Current Assets"
                            or parentgroup == "Fixed Assets"
                            or parentgroup == "Investments"
                            or parentgroup == "Loans(Asset)"
                            or parentgroup == "Miscellaneous Expenses(Asset)"
                        ):
                            openingBl = float(-openingBl)
                            try:
                                newacc = self.con.execute(
                                    accounts_table.insert(),
                                    {
                                        "accountname": accRow[0].value,
                                        "groupcode": curgrpid,
                                        "openingbal": openingBl,
                                        "orgcode": user["orgcode"],
                                    },
                                )
                            except Exception as e:
                                print(e)

                        continue
                    # checking if opening Balance is not in Credit column. i.e. column no. 2 (A).
                    # It means value is in debit column
                    if accRow[2].value == None and accRow[1].value != None:
                        openingBl = float(accRow[1].value)
                        if (
                            parentgroup == "Corpus"
                            or parentgroup == "Capital"
                            or parentgroup == "Current Liabilities"
                            or parentgroup == "Loans(Liabilities)"
                            or parentgroup == "Reserves"
                        ):
                            openingBl = float(-openingBl)
                            try:
                                newacc = self.con.execute(
                                    accounts_table.insert(),
                                    {
                                        "accountname": accRow[0].value,
                                        "groupcode": curgrpid,
                                        "openingbal": openingBl,
                                        "orgcode": user["orgcode"],
                                    },
                                )
                            except Exception as e:
                                print(e)
                            continue

                if len(accRow) == 2:
                    try:
                        newsub = self.con.execute(
                            accounts_table.insert(),
                            {
                                "accountname": accRow[0].value,
                                "groupcode": curgrpid,
                                "openingbal": accRow[1].value,
                                "orgcode": user["orgcode"],
                            },
                        )
                    except Exception as e:
                        print(e)

        # the dictionary thus returned will have
        # accountname as key and accountcode as value.
        acclist = gk_api(url="/accounts?acclist", header=header, request=self.request)
        print("account list: ", acclist["gkresult"])
        accounts = acclist["gkresult"]
        Wsheets = wbTally.worksheets
        # When data is imported from GNUKhata exported file
        if Wsheets[1].title == "Vouchers List":
            gVchList = tuple(Wsheets[1].rows)
            for gVch in gVchList:
                if gVch[1].value == None and gVch[2].value == None:
                    continue
                voucherno = gVch[0].value
                voucherdt = gVch[1].value
                vdates = voucherdt.split("-")
                voucherDt = vdates[2] + "/" + vdates[1] + "/" + vdates[0]
                vouchertype = gVch[2].value
                drs = {}
                crs = {}
                if (gVch[3].value) == "(as per details)":
                    try:
                        Vindex = gVchList.index(gVch) + 1
                        while gVchList[Vindex][3].value != None:
                            drs[accounts[gVchList[Vindex][3].value]] = gVchList[Vindex][
                                4
                            ].value
                            Vindex = Vindex + 1
                    except IndexError:
                        pass
                else:
                    drs[accounts[gVch[3].value]] = gVch[4].value

                if (gVch[5].value) == "(as per details)":
                    Vindex = gVchList.index(gVch) + 1
                    while gVchList[Vindex][5].value != None:
                        crs[accounts[gVchList[Vindex][5].value]] = gVchList[Vindex][
                            6
                        ].value
                        Vindex = Vindex + 1
                else:
                    crs[accounts[gVch[5].value]] = gVch[6].value
                    narration = gVch[7].value
                    result = {
                        "voucherdate": voucherDt,
                        "vouchernumber": gVch[0].value,
                        "vouchertype": gVch[2].value,
                        "drs": drs,
                        "crs": crs,
                        "narration": gVch[7].value,
                    }
                    gNewvch = gk_api(
                        method="POST",
                        url="/transaction",
                        body=result,
                        header=header,
                        request=self.request,
                    )
                    print(6, gNewvch)
            return {"gkstatus": 0}

        else:
            gVchList = tuple(Wsheets[1].rows)
            currentRowIndex = 0
            for gVch in gVchList:
                drs = {}
                crs = {}
                narrations = ""
                if gVch[0].value == None:
                    continue
                if gVch[0].value == "Date":
                    continue
                if gVch[0].value != None and gVch[0].value != "Date":
                    voucherdt = str(gVch[0].value)
                    voucherDt = (
                        voucherdt[0:4] + "/" + voucherdt[5:7] + "/" + voucherdt[8:10]
                    )
                    voucherType = (gVch[4].value).lower().replace(" ", "")
                    if gVch[6].value != None:
                        drs[accounts[gVch[1].value]] = str(gVch[6].value)
                    if gVch[7].value != None:
                        crs[accounts[gVch[1].value]] = str(gVch[7].value)
                        gVch = gVchList[gVchList.index(gVch) + 1]
                    while gVch[0].value == None:
                        if gVch[6].value != None:
                            drs[accounts[gVch[1].value]] = str(gVch[6].value)
                        if gVch[7].value != None:
                            crs[accounts[gVch[1].value]] = str(gVch[7].value)
                        if gVch[6].value == None and gVch[7].value == None:
                            narrations = gVch[1].value
                        try:
                            gVch = gVchList[gVchList.index(gVch) + 1]
                        except IndexError:
                            break
                        result = {
                            "voucherdate": voucherDt,
                            "vouchertype": voucherType,
                            "drs": drs,
                            "crs": crs,
                            "narration": narrations,
                        }
                        gNewvch = gk_api(
                            method="POST",
                            url="/transaction",
                            body=result,
                            header=header,
                            request=self.request,
                        )
                        print(7, gNewvch)
            return {"gkstatus": 0}
    except Exception as e:
        print(e)
        return {"gkstatus": 3}
