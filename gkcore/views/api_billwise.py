"""
Copyright (C) 2013, 2014, 2015, 2016, 2017 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
'Abhijith Balan'<abhijith@dff.org.in>
"Krishnakant Mane" <kk@dff.org.in>
"Prajkta Patkar" <prajakta@dff.org.in>
"""

from gkcore import eng, enumdict
from gkcore.views.api_login import authCheck
from gkcore.models import gkdb
from sqlalchemy.sql import select
import json
from sqlalchemy.engine.base import Connection
from sqlalchemy import and_, exc,alias, or_, func, desc
from pyramid.request import Request
from pyramid.response import Response
from pyramid.view import view_defaults, view_config
from sqlalchemy.sql.expression import null
from gkcore.models.meta import dbconnect
from gkcore.models.gkdb import billwise, invoice, customerandsupplier, vouchers,accounts,organisation
from datetime import datetime, date
from operator import itemgetter
from natsort import natsorted
import openpyxl
from openpyxl.styles import Font, Alignment
import base64
@view_defaults(route_name='billwise')
class api_billWise(object):
    """
    This class is a resource for billwise accounting.
It will be used for creating entries in the billwise table and updating it as new entries are passed.
    The invoice table will also be updated every time an adjustment is made.
    We will have get and post methods.
    """
    def __init__(self, request):
        self.request = Request
        self.request = request
        self.con = Connection
        print "billwise initialized"
    @view_config(request_method='POST',renderer='json')
    def adjustBills(self):
        """
        purpose:
        adjustment of invoices using a given receipt.
        Single receipt can be used for one or more invoices.
        description:
        this function takes a list of dictionaries containing,
        *vouchercode,
        * invid
        * amount.
        The amount key is also used to update the payed amount in invoice table.
        A for loop runs through the list of dictionaries.
        """
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                dataSet = self.request.json_body
                adjBills = dataSet["adjbills"]
                for bill in adjBills:
                    bill["orgcode"]= authDetails["orgcode"]
                    result = self.con.execute(billwise.insert(),[bill])
                    updres = self.con.execute("update invoice set amountpaid = amountpaid + %f where invid = %d"%(float(bill["adjamount"]),bill["invid"]))
                return{"gkstatus":enumdict["Success"]}
            except:
                return{"gkstatus":enumdict["ConnectionFailed"]}
    @view_config(request_method='GET',renderer='json')
    def getUnadjustedBills(self):
        """
        Purpose:
        Gets the list of unadjusted receipts and invoices.
        description:
        first we provide it a customerandsupplier code.
        Then get all the receipts vouchers for that customer or supplyer.
        now the receipts are filtered on the basis of 2 conditions.
        firstly the vouchers is unused at all.
        secondly if it is used then it's total should not be equal to the sum of all adjusted amounts of those invoices which have been adjusted using this vouchers.
        For this we loop through the list of vouchers.
        at the beginning of the loop we set a qualification flag to true.
        we will have a nested set of 2 if conditions.
        first we check if the given vouchers is present in the billwise table or not.
        if it is present then the secondly if conditions checks
        if the total amountpaid in the vouchers is equal to the sum of invoices as mentioned above.
        if this condition too is satisfied then the flag is set to false.
        if any of this conditions fail then flag remains true.
        finaly there will be a 3rd if condition which checks this flag.
        if it is true then the vouchers is added to the list to be returned.
        """
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                csid =int(self.request.params["csid"])
                csFlag =int(self.request.params["csflag"])
                csn = self.con.execute(select([customerandsupplier.c.custname]).where(and_(customerandsupplier.c.custid == csid,customerandsupplier.c.orgcode==authDetails["orgcode"])))
                csName = csn.fetchone()
                accData = self.con.execute(select([accounts.c.accountcode]).where(and_(accounts.c.accountname == csName["custname"],accounts.c.orgcode== authDetails["orgcode"])))
                acccode = accData.fetchone()
                if csFlag == 3:
                    csReceiptData = self.con.execute("select vouchercode, vouchernumber, voucherdate, crs->>'%d' as amt from vouchers where crs ? '%d' and orgcode = %d and vouchertype = 'receipt'"%(acccode["accountcode"],acccode["accountcode"],authDetails["orgcode"]))
                if csFlag == 19:
                    csReceiptData = self.con.execute("select vouchercode, vouchernumber, voucherdate, drs->>'%d' as amt from vouchers where drs ? '%d' and orgcode = %d and vouchertype = 'payment'"%(acccode["accountcode"],acccode["accountcode"],authDetails["orgcode"]))
                csReceipts = csReceiptData.fetchall()
                unAdjReceipts = []
                unAdjInvoices = []
                for rcpt in csReceipts:
                    invs = self.con.execute(select([func.count(billwise.c.invid).label('invFound'),func.sum(billwise.c.adjamount).label("amtAdjusted")]).where(billwise.c.vouchercode == rcpt["vouchercode"]))
                    invsData = invs.fetchone()
                    amtadj = 0.00
                    if int(invsData["invFound"]) > 0:
                        amtadj = invsData["amtAdjusted"]
                        
                        if float(rcpt["amt"]) == float(invsData["amtAdjusted"]):
                            continue
                    unAdjReceipts.append({"vouchercode":rcpt["vouchercode"],"vouchernumber":rcpt["vouchernumber"],"voucherdate":datetime.strftime(rcpt["voucherdate"],'%d-%m-%Y'),"amtadj":"%.2f"%(float(float(rcpt["amt"]) - float (amtadj)))})
                    
                csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid]).where(and_(invoice.c.custid == csid,invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.orgcode == authDetails["orgcode"])))
                csInvoicesData = csInvoices.fetchall()
                for inv in csInvoicesData:
                    unAdjInvoices.append({"invid":inv["invid"],"invoiceno":inv["invoiceno"],"invoicedate":datetime.strftime(inv["invoicedate"],'%d-%m-%Y'),"invoiceamount":"%.2f"%(float(inv["invoicetotal"])),"balanceamount":"%.2f"%(float(inv["invoicetotal"]-inv["amountpaid"]))})
                return{"gkstatus":enumdict["Success"],"vouchers":unAdjReceipts,"invoices":unAdjInvoices}
            except:
                return{"gkstatus":enumdict["ConnectionFailed"]}

    @view_config(request_method='GET',renderer='json', request_param="type=all")
    def getallUnadjustedBills(self):
        """
        Purpose:
        Gets the list of unadjusted invoices.
        Description:
        An invoice is considered unadjusted if it has not been paid or payment for it has not been received completely.
        These are adjusted either while creating vouchers or while doing bill wise accounting.
        This function returns a list of all unadjusted or partially adjusted bills of an organisation.
        """
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                # An empty list into which unadjusted invoices shall be appended.
                unAdjInvoices = []
                # Fetching id, number, date, total amount and amount paid of all unpaid invoices.
                # It is unadjusted if invoice total is greater that amount paid.
                invoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"])))
                invoicesData = invoices.fetchall()
                # Appending dictionaries into empty list.
                # Each dictionary has details of an invoice viz. id, number, date, total amount, amount paid and balance.
                for inv in invoicesData:
                    custData = self.con.execute(select([customerandsupplier.c.custname, customerandsupplier.c.csflag, customerandsupplier.c.custid]).where(customerandsupplier.c.custid == inv["custid"]))
                    customerdata = custData.fetchone()
                    unAdjInvoices.append({"invid":inv["invid"],"invoiceno":inv["invoiceno"],"invoicedate":datetime.strftime(inv["invoicedate"],'%d-%m-%Y'),"invoicetotal":"%.2f"%(float(inv["invoicetotal"])),"balanceamount":"%.2f"%(float(inv["invoicetotal"]-inv["amountpaid"])), "custname":customerdata["custname"], "custid":customerdata["custid"], "csflag": customerdata["csflag"]})
                return{"gkstatus":enumdict["Success"],"invoices":unAdjInvoices}
            except:
                return{"gkstatus":enumdict["ConnectionFailed"]}
    @view_config(request_method='GET',renderer='json', request_param="type=pending")
    def getallPendingBills(self):
        """
        Purpose:
        Gets the list of pending invoices.
        Description:
        An invoice is considered pending if it has not been paid or no payment for it has been received.
        These are adjusted either while creating vouchers or while doing bill wise accounting.
        This function returns a list of all pending bills of an organisation.
        """
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                # An empty list into which pending invoices shall be appended.
                unAdjInvoices = []
                # Fetching id, number, date, total amount and amount paid of all unpaid invoices.
                # It is pending if invoice total is greater that amount paid.
                invoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.amountpaid == 0, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"])))
                invoicesData = invoices.fetchall()
                # Appending dictionaries into empty list.
                # Each dictionary has details of an invoice viz. id, number, date, total amount, amount paid and balance.
                for inv in invoicesData:
                    custData = self.con.execute(select([customerandsupplier.c.custname, customerandsupplier.c.csflag, customerandsupplier.c.custid]).where(customerandsupplier.c.custid == inv["custid"]))
                    customerdata = custData.fetchone()
                    # If there is a invtype parameter then only sale/purchase invoices are returned depending on the value of type.
                    if self.request.params.has_key('invtype'):
                        if str(self.request.params["invtype"]) == 'sale' and int(customerdata['csflag']) == 3:
                            unAdjInvoices.append({"invid":inv["invid"],"invoiceno":inv["invoiceno"],"invoicedate":datetime.strftime(inv["invoicedate"],'%d-%m-%Y'),"invoicetotal":"%.2f"%(float(inv["invoicetotal"])),"balanceamount":"%.2f"%(float(inv["invoicetotal"]-inv["amountpaid"])), "custname":customerdata["custname"], "custid":customerdata["custid"], "csflag": customerdata["csflag"]})
                        elif str(self.request.params["invtype"]) == 'purchase' and int(customerdata['csflag']) == 19:
                            unAdjInvoices.append({"invid":inv["invid"],"invoiceno":inv["invoiceno"],"invoicedate":datetime.strftime(inv["invoicedate"],'%d-%m-%Y'),"invoicetotal":"%.2f"%(float(inv["invoicetotal"])),"balanceamount":"%.2f"%(float(inv["invoicetotal"]-inv["amountpaid"])), "custname":customerdata["custname"], "custid":customerdata["custid"], "csflag": customerdata["csflag"]})
                    else:
                        unAdjInvoices.append({"invid":inv["invid"],"invoiceno":inv["invoiceno"],"invoicedate":datetime.strftime(inv["invoicedate"],'%d-%m-%Y'),"invoicetotal":"%.2f"%(float(inv["invoicetotal"])),"balanceamount":"%.2f"%(float(inv["invoicetotal"]-inv["amountpaid"])), "custname":customerdata["custname"], "custid":customerdata["custid"], "csflag": customerdata["csflag"]})
                return{"gkstatus":enumdict["Success"],"invoices":unAdjInvoices}
            except:
                return{"gkstatus":enumdict["ConnectionFailed"]}

    @view_config(request_method='GET',renderer='json', request_param="type=onlybills")
    def getOnlyUnadjustedBills(self):
        """
        Purpose:
        Gets the list of invoices for a customer/supplier in the order of balance amount.
        Description:
        We receive customer id and orderflag. The value of orderflag is 1 for ascending order and 4 for descending.
        Data of invoices for the customer that are not fully paid are fetched in the order of balance amount.
        If orderflag is not 1 they are fetched in the descending order.
        A list of dictionaries is then returned where each dictionary contains data regarding an invoice.
        """
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                csid =int(self.request.params["csid"])
                orderflag = int(self.request.params["orderflag"])
                typeflag = int(self.request.params["typeflag"])
                # Period for which this report is created is determined by startdate and enddate. They are formatted as YYYY-MM-DD.
                startdate =datetime.strptime(str(self.request.params["startdate"]),"%d-%m-%Y").strftime("%Y-%m-%d")
                enddate =datetime.strptime(str(self.request.params["enddate"]),"%d-%m-%Y").strftime("%Y-%m-%d")
                # Empty list for storing incoices
                unAdjInvoices = []
                # Invoices in ascending order of amount.
                if orderflag == 1 and typeflag == 1:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid]).where(and_(invoice.c.custid == csid,invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(invoice.c.invoicetotal - invoice.c.amountpaid))
                # Invoices in descending order of amount.
                if orderflag == 4 and typeflag == 1:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid]).where(and_(invoice.c.custid == csid,invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(desc(invoice.c.invoicetotal - invoice.c.amountpaid)))
                # Invoices in ascending order of due date.
                if orderflag == 1 and typeflag == 4:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid]).where(and_(invoice.c.custid == csid,invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(invoice.c.invoicedate))
                # Invoices in descending order of due date.
                if orderflag == 4 and typeflag == 4:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid]).where(and_(invoice.c.custid == csid,invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(desc(invoice.c.invoicedate)))
                csInvoicesData = csInvoices.fetchall()
                srno = 1
                for inv in csInvoicesData:
                    unAdjInvoices.append({"invid":inv["invid"],"invoiceno":inv["invoiceno"],"invoicedate":datetime.strftime(inv["invoicedate"],'%d-%m-%Y'),"invoiceamount":"%.2f"%(float(inv["invoicetotal"])),"balanceamount":"%.2f"%(float(inv["invoicetotal"]-inv["amountpaid"])), "srno":srno})
                    srno = srno + 1
                return{"gkstatus":enumdict["Success"],"invoices":unAdjInvoices}
            except:
                return{"gkstatus":enumdict["ConnectionFailed"]}

    @view_config(request_method='GET',renderer='json', request_param="type=onlybillsforall")
    def getOnlyUnadjustedBillsForAll(self):
        """
        Purpose:
        Gets the list of invoices for all customers and suppliers in the order of balance amount and due date.
        Description:
        We receive orderflag and typeflag. The value of orderflag is 1 for ascending order and 4 for descending.
        If typeflag is 1 data of invoices for the  all customers and suppliers that are not fully paid are fetched in order of balance amount.
        If orderflag is 4 they are fetched in the descending order.
        If typeflag is 4 invoices are fetched in order of due date.
        If it is 3 invoices are fetched normally and sorted later in the order of customer/supplier name.
        This is done so because name of customer/supplier is not stored in invoice table but in customerandsupplier table.
        A list of dictionaries is then returned where each dictionary contains data regarding an invoice.
        """
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                orderflag = int(self.request.params["orderflag"])
                typeflag = int(self.request.params["typeflag"])
                # Period for which this report is created is determined by startdate and enddate. They are formatted as YYYY-MM-DD.
                startdate =datetime.strptime(str(self.request.params["startdate"]),"%d-%m-%Y").strftime("%Y-%m-%d")
                enddate =datetime.strptime(str(self.request.params["enddate"]),"%d-%m-%Y").strftime("%Y-%m-%d")
                # Empty list for storing incoices
                unAdjInvoices = []
                # Invoices in ascending order of amount.
                if orderflag == 1 and typeflag == 1:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(invoice.c.invoicetotal - invoice.c.amountpaid))
                # Invoices in descending order of amount.
                if orderflag == 4 and typeflag == 1:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(desc(invoice.c.invoicetotal - invoice.c.amountpaid)))
                # Invoices in ascending order of due date.
                if orderflag == 1 and typeflag == 4:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(invoice.c.invoicedate))
                # Invoices in descending order of due date.
                if orderflag == 4 and typeflag == 4:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(desc(invoice.c.invoicedate)))
                # Unsorted invoices to be sorted later in the order of customer/supplier name.
                if typeflag == 3:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)))
                csInvoicesData = csInvoices.fetchall()
                srno = 1
                for inv in csInvoicesData:
                    csd = self.con.execute(select([customerandsupplier.c.custname, customerandsupplier.c.csflag]).where(and_(customerandsupplier.c.custid == inv["custid"],customerandsupplier.c.orgcode==authDetails["orgcode"])))
                    csDetails = csd.fetchone()
                    unAdjInvoices.append({"invid":inv["invid"],"invoiceno":inv["invoiceno"],"invoicedate":datetime.strftime(inv["invoicedate"],'%d-%m-%Y'),"invoiceamount":"%.2f"%(float(inv["invoicetotal"])),"balanceamount":"%.2f"%(float(inv["invoicetotal"]-inv["amountpaid"])), "custname":csDetails["custname"],"csflag":csDetails["csflag"] , "srno":srno})
                    srno = srno + 1
                # List of dictionaries unAdjInvoices is sorted in order of key custname.
                if typeflag == 3 and orderflag == 1:
                    newlistofinvs = natsorted(unAdjInvoices, key=itemgetter('custname'))
                    unAdjInvoices = newlistofinvs
                if typeflag == 3 and orderflag == 4:
                    newlistofinvs = natsorted(unAdjInvoices, key=itemgetter('custname'), reverse=True)
                    unAdjInvoices = newlistofinvs
                return{"gkstatus":enumdict["Success"],"invoices":unAdjInvoices}
            except:
                return{"gkstatus":enumdict["ConnectionFailed"]}

    @view_config(request_method='GET', request_param="type=spreadsheet", renderer="json")
    def spreadsheetForReport(self):
        """
        Purpose:
        Gets the list of invoices for all customers and suppliers in the order of balance amount and due date in spreadsheet(XLSX) format.
        Description:
        For documentation regarding how data is fetched refer the function above(getOnlyUnadjustedBillsForAll).
        Documentation regarding generation of spreadsheet is given inline.
        """
        try:
            token = self.request.headers["gktoken"]
        except:
            return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"]==False:
            return {"gkstatus":enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                orderflag = int(self.request.params["orderflag"])
                typeflag = int(self.request.params["typeflag"])
                startdate =datetime.strptime(str(self.request.params["startdate"]),"%d-%m-%Y").strftime("%Y-%m-%d")
                enddate =datetime.strptime(str(self.request.params["enddate"]),"%d-%m-%Y").strftime("%Y-%m-%d")
                # A workbook is opened.
                billwisewb = openpyxl.Workbook()
                # A new sheet is created.
                billwisewb.create_sheet()
                # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
                sheet = billwisewb.active
                # Title of the sheet and width of columns are set.
                sheet.title = "List of Unpaid Invoices"
                sheet.column_dimensions['A'].width = 8
                sheet.column_dimensions['B'].width = 18
                sheet.column_dimensions['C'].width = 14
                sheet.column_dimensions['D'].width = 24
                sheet.column_dimensions['E'].width = 16
                sheet.column_dimensions['F'].width = 16
                # Cells of first two rows are merged to display organisation details properly.
                sheet.merge_cells('A1:F2')
                # Name and Financial Year of organisation is fetched to be displayed on the first row.
                orgdata = self.con.execute(select([organisation.c.orgname, organisation.c.yearstart, organisation.c.yearend]).where(organisation.c.orgcode==authDetails["orgcode"]))
                orgdetails = orgdata.fetchone()
                # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
                sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
                sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
                # Organisation name and financial year are displayed.
                sheet['A1'] = orgdetails["orgname"] + ' (FY: ' + datetime.strftime(orgdetails["yearstart"],'%d-%m-%Y') + ' to ' + datetime.strftime(orgdetails["yearend"],'%d-%m-%Y') +')'
                sheet.merge_cells('A3:F3')
                sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
                sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
                sheet['A3'] = 'List of Unpaid Invoices'
                sheet.merge_cells('A4:F4')
                sheet['A4'] = 'Period: ' + str(self.request.params["startdate"]) + ' to ' + str(self.request.params["enddate"])
                sheet['A4'].font = Font(name='Liberation Serif',size='14',bold=True)
                sheet['A4'].alignment = Alignment(horizontal = 'center', vertical='center')
                sheet['A5'] = 'Sr. No. '
                sheet['B5'] = 'Invoice No'
                sheet['C5'] = 'Invoice Date'
                sheet['D5'] = 'Cust/Supp Name'
                sheet['E5'] = 'Invoice Amount'
                sheet['F5'] = 'Amount Pending'
                titlerow = sheet.row_dimensions[5]
                titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
                sheet['E5'].alignment = Alignment(horizontal='right')
                sheet['F5'].alignment = Alignment(horizontal='right')
                sheet['E5'].font = Font(name='Liberation Serif',size=12,bold=True)
                sheet['F5'].font = Font(name='Liberation Serif',size=12,bold=True)
                # Empty list for storing incoices
                unAdjInvoices = []
                # Invoices in ascending order of amount.
                if orderflag == 1 and typeflag == 1:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(invoice.c.invoicetotal - invoice.c.amountpaid))
                # Invoices in descending order of amount.
                if orderflag == 4 and typeflag == 1:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(desc(invoice.c.invoicedate)))
                # Invoices in ascending order of due date.
                if orderflag == 1 and typeflag == 4:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(invoice.c.invoicedate))
                # Invoices in descending order of due date.
                if orderflag == 4 and typeflag == 4:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)).order_by(desc(invoice.c.invoicetotal - invoice.c.amountpaid)))
                # Unsorted invoices to be sorted later in the order of customer/supplier name.
                if typeflag == 3:
                    csInvoices = self.con.execute(select([invoice.c.invid,invoice.c.invoiceno,invoice.c.invoicedate,invoice.c.invoicetotal,invoice.c.amountpaid, invoice.c.custid]).where(and_(invoice.c.invoicetotal > invoice.c.amountpaid, invoice.c.icflag == 9, invoice.c.orgcode == authDetails["orgcode"],invoice.c.invoicedate >= startdate, invoice.c.invoicedate <= enddate)))
                csInvoicesData = csInvoices.fetchall()
                srno = 1
                for inv in csInvoicesData:
                    csd = self.con.execute(select([customerandsupplier.c.custname, customerandsupplier.c.csflag]).where(and_(customerandsupplier.c.custid == inv["custid"],customerandsupplier.c.orgcode==authDetails["orgcode"])))
                    csDetails = csd.fetchone()
                    unAdjInvoices.append({"invoiceno":inv["invoiceno"],"invoicedate":datetime.strftime(inv["invoicedate"],'%d-%m-%Y'),"invoiceamount":"%.2f"%(float(inv["invoicetotal"])),"balanceamount":"%.2f"%(float(inv["invoicetotal"]-inv["amountpaid"])), "custname":csDetails["custname"] , "srno":srno})
                    srno = srno + 1
                # List of dictionaries unAdjInvoices is sorted in order of key custname.
                if typeflag == 3 and orderflag == 1:
                    newlistofinvs = natsorted(unAdjInvoices, key=itemgetter('custname'))
                    unAdjInvoices = newlistofinvs
                if typeflag == 3 and orderflag == 4:
                    newlistofinvs = natsorted(unAdjInvoices, key=itemgetter('custname'), reverse=True)
                    unAdjInvoices = newlistofinvs
                row = 6
                # Looping each dictionaries in list unAdjInvoices to store data in cells and apply styles.
                for uninv in unAdjInvoices:
                    sheet['A'+str(row)] = uninv['srno']
                    sheet['A'+str(row)].alignment = Alignment(horizontal='left')
                    sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    sheet['B'+str(row)] = uninv['invoiceno']
                    sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    sheet['C'+str(row)] = uninv['invoicedate']
                    sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    sheet['D'+str(row)] = uninv['custname']
                    sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    sheet['E'+str(row)] = uninv['invoiceamount']
                    sheet['F'+str(row)] = uninv['balanceamount']
                    sheet['E'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['F'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    row = row + +1
                # Saving the xlsx file.
                billwisewb.save('report.xlsx')
                # Opening xlsx file in read only mode.
                reportxslx = open("report.xlsx","r")
                # Encoding xlsx file in base64 format.
                xlsxdata = base64.b64encode(reportxslx.read())
                # Closing file.
                reportxslx.close()
                return {"gkstatus":enumdict["Success"],"gkdata":xlsxdata}
            except:
                print "Spreadsheet not created."
                return {"gkstatus":3}
