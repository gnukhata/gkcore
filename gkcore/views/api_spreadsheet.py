"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Survesh" <123survesh@gmail.com>
"Sai Karthik"<kskarthik@disrot.org>

"""

from gkcore import eng, enumdict
from gkcore.views.api_login import authCheck
from gkcore.models.gkdb import (
    organisation,
    unitofmeasurement,
    product,
    goprod,
    stock,
    categorysubcategories,
)
from sqlalchemy.sql import select
from sqlalchemy.engine.base import Connection
from sqlalchemy import and_, func
from pyramid.request import Request
from pyramid.response import Response
from pyramid.view import view_defaults, view_config
import gkcore

# from gkcore.views.api_reports import getBalanceSheet
# from gkcore.views.api_invoice import getInvoiceList
from datetime import datetime
from gkcore.views.api_user import getUserRole
from gkcore.views.api_godown import getusergodowns

# Spreadsheet libraries
import io
import openpyxl
from openpyxl.styles import Font, Alignment
import gkcore.views.spreadsheets as sheets

# from openpyxl.styles.colors import RED

"""
    This API returns a spreadsheet in XLSX format of the desired report.
"""


@view_defaults(route_name="spreadsheet")
class api_spreadsheet(object):
    def __init__(self, request):
        self.request = Request
        self.request = request
        self.con = Connection
        print("Spreadsheet API initialized")

    """
    This function returns a list of products and services in the form of a spreadsheet.

    Params
        type    = products_list
        fystart = Financial start date (dd-mm-yyyy)
        fyend   = Financial end date (dd-mm-yyyy)
        orgname = Organisation name
    """

    @view_config(
        request_method="GET", request_param="type=products_list", renderer="json"
    )
    def getListofProductsSpreadsheet(self):
        try:
            token = self.request.headers["gktoken"]
        except:
            return {"gkstatus": gkcore.enumdict["UnauthorisedAccess"]}
        authDetails = authCheck(token)
        if authDetails["auth"] == False:
            return {"gkstatus": gkcore.enumdict["UnauthorisedAccess"]}
        else:
            try:
                self.con = eng.connect()
                orgdetails = self.con.execute(
                    select(
                        [
                            organisation.c.orgname,
                            organisation.c.yearstart,
                            organisation.c.yearend,
                        ]
                    ).where(organisation.c.orgcode == authDetails["orgcode"])
                )
                orgData = orgdetails.fetchone()
                gstorvatflag = 29
                date1 = "2017-07-01"
                gstdate = datetime.strptime(date1, "%Y-%m-%d")
                financialStart = datetime.strptime(
                    str(orgData["yearstart"]), "%Y-%m-%d"
                )
                financialEnd = datetime.strptime(str(orgData["yearend"]), "%Y-%m-%d")
                fystart = financialStart.strftime("%d-%m-%Y")
                fyend = financialEnd.strftime("%d-%m-%Y")
                orgname = orgData["orgname"]

                if gstdate > financialStart and gstdate > financialEnd:
                    gstorvatflag = 22
                elif gstdate > financialStart and gstdate <= financialEnd:
                    gstorvatflag = 29
                elif gstdate <= financialStart and gstdate <= financialEnd:
                    gstorvatflag = 7
                userrole = getUserRole(authDetails["userid"])
                gorole = userrole["gkresult"]
                if gorole["userrole"] == 3:
                    uId = getusergodowns(authDetails["userid"])
                    gid = []
                    for record1 in uId["gkresult"]:
                        gid.append(record1["goid"])
                    productCodes = []
                    for record2 in gid:
                        proCode = self.con.execute(
                            select([goprod.c.productcode]).where(
                                goprod.c.goid == record2
                            )
                        )
                        proCodes = proCode.fetchall()
                        for record3 in proCodes:
                            if record3["productcode"] not in productCodes:
                                productCodes.append(record3["productcode"])
                    results = []
                    for record4 in productCodes:
                        result = self.con.execute(
                            select(
                                [
                                    product.c.productcode,
                                    product.c.productdesc,
                                    product.c.categorycode,
                                    product.c.uomid,
                                    product.c.gsflag,
                                    product.c.prodsp,
                                    product.c.prodmrp,
                                ]
                            )
                            .where(
                                and_(
                                    product.c.orgcode == authDetails["orgcode"],
                                    product.c.productcode == record4,
                                )
                            )
                            .order_by(product.c.productdesc)
                        )
                        products = result.fetchone()
                        results.append(products)
                else:
                    invdc = 9
                    try:
                        invdc = int(self.request.params["invdc"])
                    except:
                        invdc = 9
                    if invdc == 4:
                        results = self.con.execute(
                            select(
                                [
                                    product.c.productcode,
                                    product.c.gsflag,
                                    product.c.productdesc,
                                    product.c.categorycode,
                                    product.c.uomid,
                                    product.c.prodsp,
                                    product.c.prodmrp,
                                ]
                            )
                            .where(
                                and_(
                                    product.c.orgcode == authDetails["orgcode"],
                                    product.c.gsflag == 7,
                                )
                            )
                            .order_by(product.c.productdesc)
                        )
                    if invdc == 9:
                        results = self.con.execute(
                            select(
                                [
                                    product.c.productcode,
                                    product.c.productdesc,
                                    product.c.gsflag,
                                    product.c.categorycode,
                                    product.c.uomid,
                                    product.c.prodsp,
                                    product.c.prodmrp,
                                ]
                            )
                            .where(product.c.orgcode == authDetails["orgcode"])
                            .order_by(product.c.productdesc)
                        )

                products = []
                srno = 1
                for productrow in results:
                    unitsofmeasurement = self.con.execute(
                        select([unitofmeasurement.c.unitname]).where(
                            unitofmeasurement.c.uomid == productrow["uomid"]
                        )
                    )
                    uom = unitsofmeasurement.fetchone()
                    if uom != None:
                        unitname = uom["unitname"]
                    else:
                        unitname = ""
                    if productrow["categorycode"] != None:
                        categories = self.con.execute(
                            select([categorysubcategories.c.categoryname]).where(
                                categorysubcategories.c.categorycode
                                == productrow["categorycode"]
                            )
                        )
                        category = categories.fetchone()
                        categoryname = category["categoryname"]
                    else:
                        categoryname = ""
                    if productrow["productcode"] != None:
                        openingStockResult = self.con.execute(
                            select([product.c.openingstock]).where(
                                product.c.productcode == productrow["productcode"]
                            )
                        )
                        osProductrow = openingStockResult.fetchone()
                        openingStock = osProductrow["openingstock"]
                        productstockin = self.con.execute(
                            select([func.sum(stock.c.qty).label("sumofins")]).where(
                                and_(
                                    stock.c.productcode == productrow["productcode"],
                                    stock.c.inout == 9,
                                )
                            )
                        )
                        stockinsum = productstockin.fetchone()
                        if stockinsum["sumofins"] != None:
                            openingStock = openingStock + stockinsum["sumofins"]
                        productstockout = self.con.execute(
                            select([func.sum(stock.c.qty).label("sumofouts")]).where(
                                and_(
                                    stock.c.productcode == productrow["productcode"],
                                    stock.c.inout == 15,
                                )
                            )
                        )
                        stockoutsum = productstockout.fetchone()
                        if stockoutsum["sumofouts"] != None:
                            openingStock = openingStock - stockoutsum["sumofouts"]
                    products.append(
                        {
                            "srno": srno,
                            "unitname": unitname,
                            "categoryname": categoryname,
                            "productcode": productrow["productcode"],
                            "productdesc": productrow["productdesc"],
                            "categorycode": productrow["categorycode"],
                            "productquantity": "%.2f" % float(openingStock),
                            "gsflag": productrow["gsflag"],
                        }
                    )
                    srno = srno + 1
                # A workbook is opened.
                productwb = openpyxl.Workbook()
                # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
                sheet = productwb.active
                # Title of the sheet and width of columns are set.
                sheet.title = "List of Products"
                sheet.column_dimensions["A"].width = 8
                sheet.column_dimensions["B"].width = 24
                sheet.column_dimensions["C"].width = 18
                sheet.column_dimensions["D"].width = 24
                sheet.column_dimensions["E"].width = 16
                # Cells of first two rows are merged to display organisation details properly.
                sheet.merge_cells("A1:E2")
                # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
                sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
                sheet["A1"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                # Organisation name and financial year are displayed.
                sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
                sheet.merge_cells("A3:E3")
                sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
                sheet["A3"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                sheet["A3"] = "List of Products"
                sheet.merge_cells("A3:E3")
                sheet["A4"] = "Sr.No."
                if gstorvatflag == "22":
                    sheet["B4"] = "Product"
                    sheet["C4"] = "Category"
                    sheet["D4"] = "UOM"
                else:
                    sheet["B4"] = "Product/service"
                    sheet["C4"] = "Type"
                    sheet["D4"] = "Category"
                    sheet["E4"] = "Uom"
                titlerow = sheet.row_dimensions[4]
                titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
                srno = 1
                if gstorvatflag == "22":
                    row = 5
                    for prodStock in products:
                        sheet["A" + str(row)] = srno
                        sheet["A" + str(row)].alignment = Alignment(horizontal="left")
                        sheet["A" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        sheet["B" + str(row)] = prodStock["productdesc"]
                        sheet["B" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        sheet["C" + str(row)] = prodStock["categoryname"]
                        sheet["C" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        sheet["D" + str(row)] = prodStock["unitname"]
                        sheet["D" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        row += 1
                        srno += 1
                else:
                    row = 5
                    for prodStock in products:
                        sheet["A" + str(row)] = srno
                        sheet["A" + str(row)].alignment = Alignment(horizontal="left")
                        sheet["A" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        sheet["B" + str(row)] = prodStock["productdesc"]
                        sheet["B" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        if prodStock["gsflag"] == 7:
                            sheet["C" + str(row)] = "Product"
                            sheet["C" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                        else:
                            sheet["C" + str(row)] = "Service"
                            sheet["C" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                        sheet["D" + str(row)] = prodStock["categoryname"]
                        sheet["D" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        sheet["E" + str(row)] = prodStock["unitname"]
                        sheet["E" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        row += 1
                        srno += 1
                output = io.BytesIO()
                productwb.save(output)
                contents = output.getvalue()
                output.close()
                headerList = {
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Length": len(contents),
                    "Content-Disposition": "attachment; filename=report.xlsx",
                    "X-Content-Type-Options": "nosniff",
                    "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
                }
                # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}
                self.con.close()
                return Response(contents, headerlist=list(headerList.items()))
            except:
                self.con.close()
                return {"gkstatus": enumdict["ConnectionFailed"]}

    @view_config(request_method="GET", request_param="stock-report", renderer="json")
    def psr(self):
        return sheets.stock_report.print_stock_report(self)

    @view_config(request_method="GET", request_param="trial-balance", renderer="json")
    def ptb(self):
        return sheets.trial_balance.print_trial_balance(self)

    @view_config(request_method="GET", request_param="pslist", renderer="json")
    def psl(self):
        return sheets.product_service.product_service_list(self)

    @view_config(request_method="GET", request_param="profit-loss", renderer="json")
    def ppl(self):
        return sheets.profit_loss.print_profit_loss(self)

    @view_config(request_method="GET", request_param="ledger", renderer="json")
    def led(self):
        return sheets.ledger.print_ledger(self)

    @view_config(request_method="GET", request_param="ledger-monthly", renderer="json")
    def ledm(self):
        return sheets.ledger.print_monthly_ledger(self)

    @view_config(
        request_method="GET", request_param="type=conv_bal_sheet", renderer="json"
    )
    def bals(self):
        return sheets.balance_sheet.print_balance_sheet(self)

    @view_config(
        request_method="GET", request_param="type=invoice-list", renderer="json"
    )
    def inv(self):
        return sheets.invoice.print_invoice_list(self)

    @view_config(request_method="GET", request_param="cash-flow", renderer="json")
    def caf(self):
        return sheets.cash_flow.print_cash_flow(self)

    @view_config(request_method="GET", request_param="accounts", renderer="json")
    def caf(self):
        return sheets.accounts.print_account_list(self)
