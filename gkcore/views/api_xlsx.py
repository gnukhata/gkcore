"""
Copyright (C) 2013, 2014, 2015, 2016, 2017 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330.
"""

from gkcore import eng, enumdict
from gkcore.views.api_login import authCheck
from gkcore.models import gkdb
from sqlalchemy.sql import select
import json
from sqlalchemy.engine.base import Connection
from sqlalchemy import and_, exc,alias, or_, func, desc
from pyramid.request import Request
from pyramid.response import Response
from pyramid.view import view_defaults, view_config
from sqlalchemy.sql.expression import null
from gkcore.models.meta import dbconnect
from gkcore.models.gkdb import  accounts, vouchers, groupsubgroups, projects, organisation, users, voucherbin,delchal,invoice,customerandsupplier,stock,product,transfernote,goprod, dcinv, log,godown, categorysubcategories, rejectionnote
from datetime import datetime, date
from monthdelta import monthdelta
from gkcore.models.meta import dbconnect
from sqlalchemy.sql.functions import func
from time import strftime, strptime
import openpyxl
from openpyxl.styles import Font, Alignment
import base64
import os

@view_defaults(route_name='worksheet')
class api_spreadsheet(object):
    """
    This class is a resource for spreadsheet.
It will be used for creating entries in the table and updating it as new entries are passed.
    We will have get and post methods.
    """
    def __init__(self, request):
        self.request = Request
        self.request = request
        self.con = Connection
    @view_config(route_name='worksheet',request_method='GET', request_param="type=sprdsheet", renderer="json")
    def spreadsheetForListOfUsers(self):
         """
         Purpose:
         Gets the list of users  in spreadsheet(XLSX) format.
         """
         try:
             token = self.request.headers["gktoken"]
         except:
             return  {"gkstatus":  enumdict["UnauthorisedAccess"]}
         authDetails = authCheck(token)
         if authDetails["auth"]==False:
             return {"gkstatus":enumdict["UnauthorisedAccess"]}
         else:
             try:
                self.con = eng.connect()
                # A workbook is opened.
                userwb = openpyxl.Workbook()
                # A new sheet is created.
                userwb.create_sheet()
                # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
                sheet = userwb.active
                # Title of the sheet and width of columns are set.
                sheet.title = "List of Users"
                sheet.column_dimensions['A'].width = 8
                sheet.column_dimensions['B'].width = 18
                sheet.column_dimensions['C'].width = 24
                sheet.column_dimensions['D'].width = 44
                # Cells of first two rows are merged to display organisation details properly.
                sheet.merge_cells('A1:G2')
                # Name and Financial Year of organisation is fetched to be displayed on the first row.
                orgdata = self.con.execute(select([organisation.c.orgname, organisation.c.yearstart, organisation.c.yearend]).where(organisation.c.orgcode==authDetails["orgcode"]))
                orgdetails = orgdata.fetchone()
                # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
                sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
                sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
                # Organisation name and financial year are displayed.
                sheet['A1'] = orgdetails["orgname"] + ' (FY: ' + datetime.strftime(orgdetails["yearstart"],'%d-%m-%Y') + ' to ' + datetime.strftime(orgdetails["yearend"],'%d-%m-%Y') +')'
                sheet.merge_cells('A3:G3')
                sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
                sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
                sheet['A3'] = 'List of Users'
                sheet.merge_cells('A3:G3')
                sheet['A4'] = 'Sr. No.'
                sheet['B4'] = 'User Name'
                sheet['C4'] = 'User Role'
                sheet['D4'] = 'Associated Godown(s)'
                titlerow = sheet.row_dimensions[4]
                titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
                result = self.con.execute(select([gkdb.users.c.username,gkdb.users.c.userid,gkdb.users.c.userrole]).where(gkdb.users.c.orgcode==authDetails["orgcode"]).order_by(gkdb.users.c.username))
                srno = 1
                row=5
                userroles={-1:"Admin",0:"Manager",1:"Operator",2:"Internal Auditor",3:"Godown In Charge"}
                #Looping to store the data in the cells and apply styles.
                for user in result:
                    urole=userroles[user["userrole"]]
                    sheet['A'+str(row)] = srno
                    sheet['A'+str(row)].alignment = Alignment(horizontal='left')
                    sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    sheet['B'+str(row)] = user['username']
                    sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    sheet['C'+str(row)] = urole
                    sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    godownresult = self.con.execute(select([gkdb.usergodown.c.goid]).where(and_(gkdb.usergodown.c.orgcode==authDetails["orgcode"], gkdb.usergodown.c.userid==user["userid"])))
                    gostring = ""
                    godownindex = 0
                    for goid in godownresult:
                            godownnameres = self.con.execute(select([gkdb.godown.c.goname, gkdb.godown.c.goaddr]).where(gkdb.godown.c.goid==goid[0]))
                            goname = godownnameres.fetchone()
                            if godownindex == 0:
                                gostring= gostring + (str(goname["goname"] + "(" +goname["goaddr"]+")"))
                            else:
                                gostring=  gostring +  ', ' + (str(goname["goname"] + "(" +goname["goaddr"]+")")) 
                            godownindex = godownindex + 1
                    if godownindex > 0:
                        sheet['D'+str(row)] = gostring
                    sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                    row = row + 1
                    srno += 1
                #Saving the xlsx file.
                userwb.save('report.xlsx')
                #Opening xlsx file in read only mode.
                reportxslx = open("report.xlsx","r")
                # Encoding xlsx file in base64 format.
                xlsxdata = base64.b64encode(reportxslx.read())
                # Closing file.
                reportxslx.close()
                os.remove("report.xlsx")
                return {"gkstatus":enumdict["Success"],"gkdata":xlsxdata}
             except:
                return {"gkstatus":3}
