"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,

Code ported from : https://gitlab.com/icfoss/texbyte_gst/
License: LGPLv3 (https://gitlab.com/icfoss/texbyte_gst/-/blob/master/texbyte_gst/LICENSE)

Contributors:
"Survesh" <123survesh@gmail.com>
"Sai Karthik"<kskarthik@disroot.org>

"""

from gkcore import eng, enumdict
from gkcore.views.api_login import authCheck
from pyramid.response import Response
import traceback  # for printing detailed exception logs

# Spreadsheet libraries
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side
from gkcore.views.api_gstreturns import generate_gstr_3b_data
from gkcore.views.api_state import getStates


def print_gstr_3b(self):

    try:
        token = self.request.headers["gktoken"]
    except:
        return {"gkstatus": enumdict["UnauthorisedAccess"]}
    authDetails = authCheck(token)
    if authDetails["auth"] == False:
        return {"gkstatus": enumdict["UnauthorisedAccess"]}
    else:
        try:
            self.con = eng.connect()
            conventionalwb = openpyxl.Workbook()
            sheet = conventionalwb.active
            sheet.title = "GSTR-3B"
            params = self.request.params

            gst_data = generate_gstr_3b_data(
                self.con,
                authDetails["orgcode"],
                params["calculatefrom"],
                params["calculateto"],
            )

            if len(gst_data):
                res = generate_3b_report(
                    self.con, self.request, conventionalwb, sheet, gst_data
                )

                if res == -1:
                    self.con.close()
                    return {"gkstatus": enumdict["ConnectionFailed"]}

                output = io.BytesIO()
                conventionalwb.save(output)
                contents = output.getvalue()
                output.close()
                headerList = {
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Length": len(contents),
                    "Content-Disposition": "attachment; filename=report.xlsx",
                    "X-Content-Type-Options": "nosniff",
                    "Set-Cookie": "fileDownload=true ;path=/ [;HttpOnly]",
                }
                # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
                self.con.close()
                return Response(contents, headerlist=list(headerList.items()))
            else:
                self.con.close()
                return {"gkstatus": enumdict["ConnectionFailed"]}
        except:
            #   print(e)
            print(traceback.format_exc())
            self.con.close()
            return {"gkstatus": enumdict["ConnectionFailed"]}


""" GSTR-3B Summary """


def generate_3b_report(con, request, wb1, ws1, gst_data):
    try:
        calculateto = request.params["calculateto"]
        calculatefrom = request.params["calculatefrom"]
        gstin = request.params["gstin"]
        orgname = request.params["orgname"]

        state_data = getStates(con)

        states = {}
        if len(state_data):
            for state in state_data:
                code = int(list(state.keys())[0])
                name = list(state.values())[0]
                states[code] = name

        ws1.column_dimensions["A"].width = 45
        ws1.column_dimensions["B"].width = 20
        ws1.column_dimensions["C"].width = 25
        ws1.column_dimensions["D"].width = 25
        ws1.column_dimensions["E"].width = 25
        ws1.column_dimensions["F"].width = 25
        ws1.column_dimensions["G"].width = 25

        sub_header_style = NamedStyle(
            name="sub_header_style",
            font=Font(name="Liberation Serif", bold=True, size="10"),
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        sub_header_content_style = NamedStyle(
            name="sub_header_content_style",
            font=Font(name="Liberation Serif", size="10"),
        )
        line_content_style = NamedStyle(
            name="line_content_style", font=Font(name="Liberation Serif")
        )

        wb1.add_named_style(sub_header_style)
        wb1.add_named_style(sub_header_content_style)
        wb1.add_named_style(line_content_style)

        # header_content_style = xlwt.easyxf("font: name Liberation Serif size 12 px, bold 1, height 170;")
        # sub_header_style = xlwt.easyxf("font: name Liberation Serif size 10 px, bold 1, height 170; align: horiz center")
        # sub_header_content_style = xlwt.easyxf("font: name Liberation Serif size 10 px, height 170;")
        # line_content_style = xlwt.easyxf("font: name Liberation Serif, height 170;")

        row = 1
        col = 0
        # ws1.row(row).height = 500
        ws1.merge_cells("A1:E1")
        ws1["A1"].font = Font(name="Liberation Serif", size="12", bold=True)
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1["A1"] = "GSTR-3B"
        row += 2

        cellRef = ws1.cell(column=(col + 1), row=row, value="From:")
        cellRef.style = "sub_header_style"
        cellRef = ws1.cell(column=(col + 2), row=row, value=calculatefrom)
        cellRef.style = "sub_header_content_style"
        row += 1

        cellRef = ws1.cell(column=(col + 1), row=row, value="To:")
        cellRef.style = "sub_header_style"
        cellRef = ws1.cell(column=(col + 2), row=row, value=calculateto)
        cellRef.style = "sub_header_content_style"
        row += 1

        cellRef = ws1.cell(column=(col + 1), row=row, value="GSTIN")
        cellRef.style = "sub_header_style"
        cellRef = ws1.cell(column=(col + 2), row=row, value=gstin)
        cellRef.style = "sub_header_content_style"
        row += 1

        cellRef = ws1.cell(
            column=(col + 1), row=row, value="Legal name of the registered person"
        )
        cellRef.style = "sub_header_style"
        cellRef = ws1.cell(column=(col + 2), row=row, value=orgname)
        cellRef.style = "sub_header_content_style"

        hsn_summary_data = {}
        # print(gst_data)
        outward_taxable_supplies = gst_data["outward_taxable_supplies"]
        outward_taxable_zero_rated = gst_data["outward_taxable_zero_rated"]
        outward_taxable_exempted = gst_data["outward_taxable_exempted"]
        outward_non_gst = gst_data["outward_non_gst"]
        inward_reverse_charge = gst_data["inward_reverse_charge"]
        import_goods = gst_data["import_goods"]
        import_service = gst_data["import_service"]
        inward_isd = gst_data["inward_isd"]
        all_itc = gst_data["all_itc"]
        itc_reversed_1 = gst_data["itc_reversed_1"]
        itc_reversed_2 = gst_data["itc_reversed_2"]
        net_itc = gst_data["net_itc"]
        ineligible_1 = gst_data["ineligible_1"]
        ineligible_2 = gst_data["ineligible_2"]
        inward_zero_gst = gst_data["inward_zero_gst"]
        non_gst = gst_data["non_gst"]
        interest = gst_data["interest"]
        pos_unreg_comp_uin_igst = gst_data["pos_unreg_comp_uin_igst"]
        row += 2

        # Innter functions
        def prepare_outward_supplies(row):
            cellRef = ws1.cell(
                column=(col + 1),
                row=row,
                value="3.1 Details of Outward Supplies and inward supplies liable to reverse charge",
            )
            cellRef.style = "sub_header_style"
            ws1.merge_cells(
                start_row=row, start_column=col + 1, end_row=row, end_column=col + 6
            )

            row += 1

            cellRef = ws1.cell(column=col + 1, row=row, value="Nature of Supplies")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 2, row=row, value="Taxable Value")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 3, row=row, value="IGST")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 4, row=row, value="CGST")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 5, row=row, value="SGST")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 6, row=row, value="Cess")
            cellRef.style = "sub_header_style"

            cellRef = ws1.cell(
                column=col + 1,
                row=row + 1,
                value="(a) Outward Taxable  supplies  (other than zero rated, nil rated and exempted)",
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1,
                row=row + 2,
                value="(b) Outward Taxable  supplies  (zero rated )",
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1,
                row=row + 3,
                value="(c) Other Outward Taxable  supplies (Nil rated, exempted)",
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1,
                row=row + 4,
                value="(d) Inward supplies (liable to reverse charge)",
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1, row=row + 5, value="(e) Non-GST Outward supplies"
            )
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(
                column=col + 2,
                row=row + 1,
                value=outward_taxable_supplies["taxable_value"],
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 2,
                row=row + 2,
                value=outward_taxable_zero_rated["taxable_value"],
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 2,
                row=row + 3,
                value=outward_taxable_exempted["taxable_value"],
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 2,
                row=row + 4,
                value=inward_reverse_charge["taxable_value"],
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 2, row=row + 5, value=outward_non_gst["taxable_value"]
            )
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(
                column=col + 3, row=row + 1, value=outward_taxable_supplies["igst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 2, value=outward_taxable_zero_rated["igst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 3, value=outward_taxable_exempted["igst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 4, value=inward_reverse_charge["igst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 5, value=outward_non_gst["igst"]
            )
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(
                column=col + 4, row=row + 1, value=outward_taxable_supplies["cgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 4, row=row + 2, value=outward_taxable_zero_rated["cgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 4, row=row + 3, value=outward_taxable_exempted["cgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 4, row=row + 4, value=inward_reverse_charge["cgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 4, row=row + 5, value=outward_non_gst["cgst"]
            )
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(
                column=col + 5, row=row + 1, value=outward_taxable_supplies["sgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 2, value=outward_taxable_zero_rated["sgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 3, value=outward_taxable_exempted["sgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 4, value=inward_reverse_charge["sgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 5, value=outward_non_gst["sgst"]
            )
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(
                column=col + 6, row=row + 1, value=outward_taxable_supplies["cess"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 6, row=row + 2, value=outward_taxable_zero_rated["cess"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 6, row=row + 3, value=outward_taxable_exempted["cess"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 6, row=row + 4, value=inward_reverse_charge["cess"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 6, row=row + 5, value=outward_non_gst["cess"]
            )
            cellRef.style = "line_content_style"

            row += 8
            return row

        def prepare_eligible_itc(row):
            cellRef = ws1.cell(column=(col + 1), row=row, value="4. Eligible ITC")
            cellRef.style = "sub_header_style"
            ws1.merge_cells(
                start_row=row, start_column=col + 1, end_row=row, end_column=col + 5
            )

            row += 1

            cellRef = ws1.cell(column=col + 1, row=row, value="Details")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 2, row=row, value="Integrated Tax")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 3, row=row, value="Central Tax")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 4, row=row, value="State/UT Tax")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 5, row=row, value="CESS")
            cellRef.style = "sub_header_style"

            cellRef = ws1.cell(
                column=col + 1,
                row=row + 1,
                value="(A) ITC Available (Whether in full or part)",
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1, row=row + 2, value="   (1) Import of goods"
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1, row=row + 3, value="   (2) Import of services"
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1,
                row=row + 4,
                value="   (3) Inward supplies liable to reverse charge(other than 1 &2 above)",
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1, row=row + 5, value="   (4) Inward supplies from ISD"
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1, row=row + 6, value="   (5) All other ITC"
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 1, row=row + 7, value="(B) ITC Reversed")
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1,
                row=row + 8,
                value="   (1) As per Rule 42 & 43 of SGST/CGST rules",
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 1, row=row + 9, value="   (2) Others")
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1, row=row + 10, value="(C) Net ITC Available (A)-(B)"
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 1, row=row + 11, value="(D) Ineligible ITC")
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 1,
                row=row + 12,
                value="  (1) As per section 17(5) of CGST/SGST Act",
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 1, row=row + 13, value="  (2) Others")
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(column=col + 2, row=row + 2, value=import_goods["igst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 2, row=row + 3, value=import_service["igst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 2, row=row + 4, value=inward_reverse_charge["igst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 2, row=row + 5, value=inward_isd["igst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 2, row=row + 6, value=all_itc["igst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 2, row=row + 8, value=itc_reversed_1["igst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 2, row=row + 9, value=itc_reversed_2["igst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 2, row=row + 10, value=net_itc["igst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 2, row=row + 12, value=ineligible_1["igst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 2, row=row + 13, value=ineligible_2["igst"])
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(column=col + 3, row=row + 2, value=import_goods["cgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 3, value=import_service["cgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 4, value=inward_reverse_charge["cgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 3, row=row + 5, value=inward_isd["cgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 3, row=row + 6, value=all_itc["cgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 8, value=itc_reversed_1["cgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 9, value=itc_reversed_2["cgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 10, value=net_itc["cgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 3, row=row + 12, value=ineligible_1["cgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 3, row=row + 13, value=ineligible_2["cgst"])
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(column=col + 4, row=row + 2, value=import_goods["sgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 4, row=row + 3, value=import_service["sgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 4, row=row + 4, value=inward_reverse_charge["sgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 4, row=row + 5, value=inward_isd["sgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 4, row=row + 6, value=all_itc["sgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 4, row=row + 8, value=itc_reversed_1["sgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 4, row=row + 9, value=itc_reversed_2["sgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 4, row=row + 10, value=net_itc["sgst"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 4, row=row + 12, value=ineligible_1["sgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 4, row=row + 13, value=ineligible_2["sgst"])
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(column=col + 5, row=row + 2, value=import_goods["cess"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 3, value=import_service["cess"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 4, value=inward_reverse_charge["cess"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 5, row=row + 5, value=inward_isd["cess"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 5, row=row + 6, value=all_itc["cess"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 8, value=itc_reversed_1["cess"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 9, value=itc_reversed_2["cess"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 10, value=net_itc["cess"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 5, row=row + 12, value=ineligible_1["cess"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 5, row=row + 13, value=ineligible_2["cess"])
            cellRef.style = "line_content_style"

            row += 16
            return row

        def prepare_exempt_supplies(row):
            cellRef = ws1.cell(
                column=(col + 1),
                row=row,
                value="5. Values of exempt, Nil-rated and non-GST inward supplies",
            )
            cellRef.style = "sub_header_style"
            ws1.merge_cells(
                start_row=row, start_column=col + 1, end_row=row, end_column=col + 3
            )

            row += 1

            cellRef = ws1.cell(column=col + 1, row=row, value="Nature of supplies")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 2, row=row, value="Inter-State Supplies")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 3, row=row, value="Intra-State Supplies")
            cellRef.style = "sub_header_style"

            cellRef = ws1.cell(
                column=col + 1,
                row=row + 1,
                value="From a supplier under composition scheme, Exempt  and Nil rated supply",
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 1, row=row + 2, value="Non-GST Supply")
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(
                column=col + 2, row=row + 1, value=inward_zero_gst["inter"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 2, row=row + 2, value=non_gst["inter"])
            cellRef.style = "line_content_style"

            cellRef = ws1.cell(
                column=col + 3, row=row + 1, value=inward_zero_gst["intra"]
            )
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 3, row=row + 2, value=non_gst["intra"])
            cellRef.style = "line_content_style"

            row += 5
            return row

        def prepare_interest_late_fee(row):
            cellRef = ws1.cell(
                column=(col + 1), row=row, value="5.1 Interest & late fee payable"
            )
            cellRef.style = "sub_header_style"
            ws1.merge_cells(
                start_row=row, start_column=col + 1, end_row=row, end_column=col + 5
            )

            row += 1

            cellRef = ws1.cell(column=col + 1, row=row, value="Description")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 2, row=row, value="Integrated Tax")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 3, row=row, value="Central Tax")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 4, row=row, value="State/UT Tax")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 5, row=row, value="CESS")
            cellRef.style = "sub_header_style"

            cellRef = ws1.cell(column=col + 1, row=row + 1, value="Interest")
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 2, row=row + 1, value=interest["igst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 3, row=row + 1, value=interest["cgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 4, row=row + 1, value=interest["sgst"])
            cellRef.style = "line_content_style"
            cellRef = ws1.cell(column=col + 5, row=row + 1, value=interest["cess"])
            cellRef.style = "line_content_style"

            row += 4
            return row

        def prepare_inter_state_unreg(row):
            cellRef = ws1.cell(
                column=(col + 1),
                row=row,
                value="3.2  Of the supplies shown in 3.1 (a), details of inter-state supplies made to unregistered persons, composition taxable person and UIN holders",
            )
            cellRef.style = "sub_header_style"
            ws1.merge_cells(
                start_row=row, start_column=col + 1, end_row=row, end_column=col + 7
            )

            row += 1

            cellRef = ws1.cell(
                column=(col + 1), row=row, value="Place of Supply(State/UT)"
            )
            cellRef.style = "sub_header_style"
            ws1.merge_cells(
                start_row=row, start_column=col + 1, end_row=row + 1, end_column=col + 1
            )

            cellRef = ws1.cell(
                column=(col + 2), row=row, value="Supplies made to Unregistered Persons"
            )
            cellRef.style = "sub_header_style"
            ws1.merge_cells(
                start_row=row, start_column=col + 2, end_row=row, end_column=col + 3
            )

            cellRef = ws1.cell(
                column=(col + 4),
                row=row,
                value="Supplies made to Composition Taxable Persons",
            )
            cellRef.style = "sub_header_style"
            ws1.merge_cells(
                start_row=row, start_column=col + 4, end_row=row, end_column=col + 5
            )

            cellRef = ws1.cell(
                column=(col + 6), row=row, value="Supplies made to UIN holders"
            )
            cellRef.style = "sub_header_style"
            ws1.merge_cells(
                start_row=row, start_column=col + 6, end_row=row, end_column=col + 7
            )

            cellRef = ws1.cell(column=col + 2, row=row + 1, value="Total Taxable value")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(
                column=col + 3, row=row + 1, value="Amount of Integrated Tax"
            )
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 4, row=row + 1, value="Total Taxable value")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(
                column=col + 5, row=row + 1, value="Amount of Integrated Tax"
            )
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 6, row=row + 1, value="Total Taxable value")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(
                column=col + 7, row=row + 1, value="Amount of Integrated Tax"
            )
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 1, row=row + 2, value="1")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 2, row=row + 2, value="2")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 3, row=row + 2, value="3")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 4, row=row + 2, value="4")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 5, row=row + 2, value="5")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 6, row=row + 2, value="6")
            cellRef.style = "sub_header_style"
            cellRef = ws1.cell(column=col + 7, row=row + 2, value="7")
            cellRef.style = "sub_header_style"

            row += 2
            for place_of_supply, tx_line in pos_unreg_comp_uin_igst.items():
                row += 1
                state_name = states[int(place_of_supply)] or place_of_supply
                cellRef = ws1.cell(column=col + 1, row=row, value=state_name)
                cellRef.style = "line_content_style"
                cellRef = ws1.cell(
                    column=col + 2, row=row, value=tx_line["unreg_taxable_amt"]
                )
                cellRef.style = "line_content_style"
                cellRef = ws1.cell(column=col + 3, row=row, value=tx_line["unreg_igst"])
                cellRef.style = "line_content_style"
                cellRef = ws1.cell(
                    column=col + 4, row=row, value=tx_line["comp_taxable_amt"]
                )
                cellRef.style = "line_content_style"
                cellRef = ws1.cell(column=col + 5, row=row, value=tx_line["comp_igst"])
                cellRef.style = "line_content_style"
                cellRef = ws1.cell(
                    column=col + 6, row=row, value=tx_line["uin_taxable_amt"]
                )
                cellRef.style = "line_content_style"
                cellRef = ws1.cell(column=col + 7, row=row, value=tx_line["uin_igst"])
                cellRef.style = "line_content_style"

            return row

        # Call the inner functions
        row = prepare_outward_supplies(row)
        row = prepare_eligible_itc(row)
        row = prepare_exempt_supplies(row)
        row = prepare_interest_late_fee(row)
        row = prepare_inter_state_unreg(row)
    except:
        print(traceback.format_exc())
        return -1
