"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020, 2021 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors
============
Survesh VRL <123survesh@gmail.com>
Sai Karthik <kskarthik@disroot.org>

"""
from gkcore.models.meta import gk_api

from pyramid.response import Response
from datetime import datetime

# Spreadsheet libraries
import io
import openpyxl
from openpyxl.styles import Font, Alignment

# from openpyxl.styles.colors import RED


def unbilled(self):
    """Unbilled delivery challans
    params:
    ======
    orgname: string
    fystart: dd-mm-yyyy
    fyend: dd-mm-yyyy
    inputdate:
    del_unbillled_type: (Values: All | Approval | Consignment | Sale | Purchase)
    inout: (Values: 9, 15)
    """
    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        orgname = str(self.request.params["orgname"])
        # orgname += " (FY: " + fystart+" to "+fyend +")"
        inputdate = self.request.params["inputdate"]
        del_unbilled_type = str(self.request.params["del_unbilled_type"])
        # financial_start = self.request.params["financial_start"]
        deltype = ""
        if del_unbilled_type == "All":
            del_unbilled_type = "0"
            deltype = "All Types"
            # merge = 6
        elif del_unbilled_type == "Approval":
            del_unbilled_type = "1"
            deltype = "Delivery Type : Approval"
            # merge = 5
        elif del_unbilled_type == "Consignment":
            del_unbilled_type = "3"
            deltype = "Delivery Type : Consignment"
            # merge = 5
        elif del_unbilled_type == "Sale":
            del_unbilled_type = "4"
            deltype = "Delivery Type : Sale"
            # merge = 5
        elif del_unbilled_type == "Purchase":
            del_unbilled_type = "16"
            deltype = "Delivery Type : Purchase"
            # merge = 5
        # gkdata = {"inputdate": inputdate, "del_unbilled_type": del_unbilled_type}
        new_inputdate = datetime.strftime(
            datetime.strptime(str(inputdate), "%Y-%m-%d").date(), "%d-%m-%Y"
        )
        inout = self.request.params["inout"]
        if inout == "9":
            result = gk_api(
                f"/report?type=del_unbilled&inout=i&inputdate={inputdate}&del_unbilled_type={del_unbilled_type}",
                header,
                self.request,
            )["gkresult"]
            headingtext = (
                "Inward Deliveries - Invoices Not Received | All Godowns | %s" % deltype
            )
            title = "Supplier Name"
        elif inout == "15":
            result = gk_api(
                "/report?type=del_unbilled&inout=o&inputdate={inputdate}&del_unbilled_type={del_unbilled_type}",
                header,
                self.request,
            )["gkresult"]
            headingtext = (
                "Outward Deliveries - Invoices Not Received | All Godowns | %s"
                % deltype
            )
            title = "Customer Name"
        # result = result.json()["gkresult"]
        # A workbook is opened.
        unbilldelwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = unbilldelwb.active
        # Title of the sheet and width of columns are set
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 16
        sheet.column_dimensions["F"].width = 16
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells("A1:F2")
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet.merge_cells("A3:F3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells("A4:F4")
        sheet["A4"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
        if inout == "9":
            sheet.title = "Deliveries In"
        elif inout == "15":
            sheet.title = "Deliveries Out"
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet["A3"] = headingtext
        sheet["A4"] = "As on Date: " + new_inputdate
        if self.request.params["del_unbilled_type"] == "9":
            sheet.title = "Deliveries In"
            sheet["A4"] = "Deliveries In"
        elif self.request.params["del_unbilled_type"] == "15":
            sheet.title = "Deliveries Out"
            sheet["A4"] = "Deliveries Out"
        sheet["A5"] = "Sr. No."
        sheet["B5"] = "Deli. Note No."
        sheet["C5"] = "Deli. Note Date"
        sheet["D5"] = title
        sheet["E5"] = "Godown Name"
        if del_unbilled_type == "0":
            sheet["F5"] = "Delivery Type"
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
        titlerow.alignment = Alignment(horizontal="center", vertical="center")
        row = 6
        for deliverychallan in result:
            sheet["A" + str(row)] = deliverychallan["srno"]
            sheet["A" + str(row)].alignment = Alignment(horizontal="center")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["B" + str(row)] = deliverychallan["dcno"]
            sheet["B" + str(row)].alignment = Alignment(horizontal="center")
            sheet["B" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["C" + str(row)] = deliverychallan["dcdate"]
            sheet["C" + str(row)].alignment = Alignment(horizontal="center")
            sheet["C" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["D" + str(row)] = deliverychallan["custname"]
            sheet["D" + str(row)].alignment = Alignment(horizontal="center")
            sheet["D" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["E" + str(row)] = deliverychallan["goname"]
            sheet["E" + str(row)].alignment = Alignment(horizontal="center")
            sheet["E" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            if del_unbilled_type == "0":
                sheet["F" + str(row)] = deliverychallan["dcflag"]
                sheet["F" + str(row)].alignment = Alignment(horizontal="center")
                sheet["F" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            row += 1
        output = io.BytesIO()
        unbilldelwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Cotent-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print("delivery challan unbilled error")
        print(e)
        return {"gkstatus": 3}


def cancelled(self):
    """Cancelled delivery challans as spreadsheet

    params:
    ======
    fystart: dd-mm-yyyy
    fyend: dd-mm-yyyy
    inputdate:
    del_cancelled_type: (Values: All | Approval | Consignment | Sale | Purchase)
    inout: (Values: 9, 15)
    """

    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        orgname = str(self.request.params["orgname"])
        # orgname += " (FY: " + fystart+" to "+fyend +")"
        inputdate = self.request.params["inputdate"]
        del_cancelled_type = str(self.request.params["del_cancelled_type"])
        deltype = ""
        if del_cancelled_type == "All":
            del_cancelled_type = "0"
            deltype = "All Types"
        elif del_cancelled_type == "Approval":
            del_cancelled_type = "1"
            deltype = "Delivery Type : Approval"
        elif del_cancelled_type == "Consignment":
            del_cancelled_type = "3"
            deltype = "Delivery Type : Consignment"
        elif del_cancelled_type == "Sale":
            del_cancelled_type = "4"
            deltype = "Delivery Type : Sale"
        elif del_cancelled_type == "Purchase":
            del_cancelled_type = "16"
            deltype = "Delivery Type : Purchase"
        # gkdata = {"inputdate": inputdate, "del_cancelled_type": del_cancelled_type}
        new_inputdate = datetime.strftime(
            datetime.strptime(str(inputdate), "%Y-%m-%d").date(), "%d-%m-%Y"
        )
        inout = self.request.params["inout"]
        if inout == "9":
            result = gk_api(
                f"/delchal?type=listofcancelleddel&inout=i&inputdate={inputdate}&del_cancelled_type={del_cancelled_type}",
                header,
                self.request,
            )["gkresult"]
            headingtext = (
                "Cancelled Inward Deliveries - Invoices Not Received | All Godowns | %s"
                % deltype
            )
            title = "Supplier Name"
        elif inout == "15":
            result = gk_api(
                f"/delchal?type=listofcancelleddel&inout=o&inputdate={inputdate}&del_cancelled_type={del_cancelled_type}",
                header,
                self.request,
            )["gkresult"]
            headingtext = (
                "Cancelled Outward Deliveries - Invoices Not Received | All Godowns | %s"
                % deltype
            )
            title = "Customer Name"
        # result = result.json()["gkresult"]
        # A workbook is opened.
        unbilldelwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = unbilldelwb.active
        # Title of the sheet and width of columns are set
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 16
        sheet.column_dimensions["F"].width = 16
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells("A1:F2")
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet.merge_cells("A3:F3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells("A4:F4")
        sheet["A4"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
        if inout == "9":
            sheet.title = "Cancelled Deliveries In"
        elif inout == "15":
            sheet.title = "Cancelled Deliveries Out"
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet["A3"] = headingtext
        sheet["A4"] = "As on Date: " + new_inputdate
        if self.request.params["del_cancelled_type"] == "9":
            sheet.title = "Cancelled Deliveries In"
            sheet["A4"] = "Cancelled Deliveries In"
        elif self.request.params["del_cancelled_type"] == "15":
            sheet.title = "Cancelled Deliveries Out"
            sheet["A4"] = "Cancelled Deliveries Out"
        sheet["A5"] = "Sr. No."
        sheet["B5"] = "Deli. Note No."
        sheet["C5"] = "Deli. Note Date"
        sheet["D5"] = title
        sheet["E5"] = "Godown Name"
        if del_cancelled_type == "0":
            sheet["F5"] = "Delivery Type"
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
        titlerow.alignment = Alignment(horizontal="center", vertical="center")
        row = 6
        for deliverychallan in result:
            sheet["A" + str(row)] = deliverychallan["srno"]
            sheet["A" + str(row)].alignment = Alignment(horizontal="center")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["B" + str(row)] = deliverychallan["dcno"]
            sheet["B" + str(row)].alignment = Alignment(horizontal="center")
            sheet["B" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["C" + str(row)] = deliverychallan["dcdate"]
            sheet["C" + str(row)].alignment = Alignment(horizontal="center")
            sheet["C" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["D" + str(row)] = deliverychallan["custname"]
            sheet["D" + str(row)].alignment = Alignment(horizontal="center")
            sheet["D" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["E" + str(row)] = deliverychallan["goname"]
            sheet["E" + str(row)].alignment = Alignment(horizontal="center")
            sheet["E" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            if del_cancelled_type == "0":
                sheet["F" + str(row)] = deliverychallan["dcflag"]
                sheet["F" + str(row)].alignment = Alignment(horizontal="center")
                sheet["F" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            row += 1
        output = io.BytesIO()
        unbilldelwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Cotent-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print(e)
        return {"gkstatus": 3}
