"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020, 2021 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors
============
Survesh VRL <123survesh@gmail.com>
Sai Karthik <kskarthik@disroot.org>

"""
from gkcore import eng, enumdict
from gkcore.models.meta import gk_api
from gkcore.views.api_login import authCheck
from pyramid.response import Response


from gkcore.views.api_invoice import getInvoiceList
from datetime import datetime

# Spreadsheet libraries
import io
import openpyxl
from openpyxl.styles import Font, Alignment

# from openpyxl.styles.colors import RED


def invoice_list(self):
    """
    This function returns a list of invoices in the form of a spreadsheet.

    Params
        type     = invoice_list
        flag     = 0 - all invoices, 1 - sale invoices, 2 - purchase invoices
        fromdate = Choose invoices from date (yyyy-mm-dd)
        todate   = Choose invoices to date (yyyy-mm-dd)
        fystart  = Financial start date (dd-mm-yyyy)
        fyend    = Financial end date (dd-mm-yyyy)
        orgname  = Organisation name
    """
    try:
        token = self.request.headers["gktoken"]
    except:
        return {"gkstatus": enumdict["UnauthorisedAccess"]}
    authDetails = authCheck(token)
    if authDetails["auth"] == False:
        return {"gkstatus": enumdict["UnauthorisedAccess"]}
    else:
        try:
            self.con = eng.connect()

            # params fromdate and todate are required to fetch the invoices list
            fystart = str(self.request.params["fystart"])
            fyend = str(self.request.params["fyend"])
            orgname = str(self.request.params["orgname"])
            invflag = int(self.request.params["flag"])
            invoices = getInvoiceList(
                self.con, authDetails["orgcode"], self.request.params
            )
            invoicewb = openpyxl.Workbook()
            sheet = invoicewb.active
            sheet.column_dimensions["A"].width = 8
            sheet.column_dimensions["B"].width = 12
            sheet.column_dimensions["C"].width = 10
            sheet.column_dimensions["D"].width = 16
            sheet.column_dimensions["E"].width = 16
            sheet.column_dimensions["F"].width = 16
            sheet.column_dimensions["G"].width = 16
            sheet.column_dimensions["H"].width = 10
            sheet.column_dimensions["I"].width = 16
            sheet.merge_cells("A1:K2")
            sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
            sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
            sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
            sheet.merge_cells("A3:K3")
            sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
            sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
            if invflag == 0:
                sheet.title = "List of All Invoices"
                sheet["A3"] = "List of All Invoices"
            elif invflag == 1:
                sheet.title = "List of Sales Invoices"
                sheet["A3"] = "List of Sales Invoices"
            elif invflag == 2:
                sheet.title = "List of Purchase Invoices"
                sheet["A3"] = "List of Purchase Invoices"
            sheet.merge_cells("A4:K4")
            sheet["A4"] = (
                "Period: "
                + datetime.strptime(
                    self.request.params["fromdate"], "%Y-%m-%d"
                ).strftime("%d-%m-%Y")
                + " to "
                + datetime.strptime(self.request.params["todate"], "%Y-%m-%d").strftime(
                    "%d-%m-%Y"
                )
            )

            sheet["A4"].font = Font(name="Liberation Serif", size="14", bold=True)
            sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
            sheet["A5"] = "Sr. No."
            sheet["B5"] = "INV No."
            sheet["C5"] = "INV Date"
            sheet["D5"] = "Deli. Note "
            if invflag == 0:
                sheet["E5"] = "Cust/Supp Name"
            elif invflag == 1:
                sheet["E5"] = "Customer Name"
            elif invflag == 2:
                sheet["E5"] = "Supplier Name"
            sheet["F5"] = "Gross Amt"
            sheet["G5"] = "Net Amt"
            sheet["H5"] = "Tax Amt"
            sheet["I5"] = "Godown"
            titlerow = sheet.row_dimensions[5]
            titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["A5"].alignment = Alignment(horizontal="center")
            sheet["B5"].alignment = Alignment(horizontal="center")
            sheet["C5"].alignment = Alignment(horizontal="center")
            sheet["D5"].alignment = Alignment(horizontal="center")
            sheet["E5"].alignment = Alignment(horizontal="center")
            sheet["F5"].alignment = Alignment(horizontal="right")
            sheet["G5"].alignment = Alignment(horizontal="right")
            sheet["H5"].alignment = Alignment(horizontal="right")
            sheet["I5"].alignment = Alignment(horizontal="center")
            sheet["A5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["B5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["C5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["D5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["E5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["F5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["G5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["H5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["I5"].font = Font(name="Liberation Serif", size=12, bold=True)
            row = 6
            # Looping each dictionaries in list result to store data in cells and apply styles.
            for sheetdata in invoices:
                sheet["A" + str(row)] = sheetdata["srno"]
                sheet["A" + str(row)].alignment = Alignment(horizontal="center")
                sheet["A" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["B" + str(row)] = sheetdata["invoiceno"]
                sheet["B" + str(row)].alignment = Alignment(horizontal="center")
                sheet["B" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["C" + str(row)] = sheetdata["invoicedate"]
                sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                sheet["C" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                if sheetdata["dcno"] != "" and sheetdata["dcdate"] != "":
                    sheet["D" + str(row)] = (
                        sheetdata["dcno"] + "," + sheetdata["dcdate"]
                    )
                    sheet["D" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                sheet["E" + str(row)] = sheetdata["custname"]
                sheet["E" + str(row)].alignment = Alignment(horizontal="center")
                sheet["E" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["F" + str(row)] = float("%.2f" % float(sheetdata["grossamt"]))
                sheet["F" + str(row)].number_format = "0.00"
                sheet["F" + str(row)].alignment = Alignment(horizontal="right")
                sheet["F" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["G" + str(row)] = float("%.2f" % float(sheetdata["netamt"]))
                sheet["G" + str(row)].number_format = "0.00"
                sheet["G" + str(row)].alignment = Alignment(horizontal="right")
                sheet["G" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["H" + str(row)] = float("%.2f" % float(sheetdata["taxamt"]))
                sheet["H" + str(row)].number_format = "0.00"
                sheet["H" + str(row)].alignment = Alignment(horizontal="right")
                sheet["H" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["I" + str(row)] = sheetdata["godown"]
                sheet["I" + str(row)].alignment = Alignment(horizontal="center")
                sheet["I" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                row = row + 1
            output = io.BytesIO()
            invoicewb.save(output)
            contents = output.getvalue()
            output.close()
            headerList = {
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Length": len(contents),
                "Content-Disposition": "attachment; filename=report.xlsx",
                "X-Content-Type-Options": "nosniff",
                "Set-Cookie": "fileDownload=true ;path=/ [;HttpOnly]",
            }
            # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
            self.con.close()
            return Response(contents, headerlist=list(headerList.items()))
        except:
            self.con.close()
            return {"gkstatus": enumdict["ConnectionFailed"]}


def cancelled_invoices(self):
    """
    This function returns a list of cancelled invoices in the form of a spreadsheet.

    Params
        invflag     = 0 - all invoices, 1 - sale invoices, 2 - purchase invoices
        fromdate   = Choose invoices from date (yyyy-mm-dd)
        todate   = Choose invoices to date (yyyy-mm-dd)
        fystart  = Financial start date (dd-mm-yyyy)
        fyend    = Financial end date (dd-mm-yyyy)
        orgname  = Organisation name
    """

    try:
        header = {"gktoken": self.request.headers["gktoken"]}

        # resultgstvat = gk_api("/products?tax=vatorgst", header, self.request)
        # resultgstvat = resultgstvat["gkresult"]
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        orgname = str(self.request.params["orgname"])
        invflag = int(self.request.params["flag"])
        from_date = self.request.params["fromdate"]
        to_date = self.request.params["todate"]

        result = gk_api(
            f"/invoice?type=listdeleted&flag={invflag}&fromdate={from_date}&todate={to_date}",
            header,
            self.request,
        )["gkresult"]
        invoicewb = openpyxl.Workbook()
        sheet = invoicewb.active
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 16
        sheet.column_dimensions["E"].width = 16
        sheet.column_dimensions["F"].width = 16
        sheet.column_dimensions["G"].width = 16
        sheet.column_dimensions["H"].width = 10
        sheet.column_dimensions["I"].width = 16
        sheet.merge_cells("A1:K2")
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet.merge_cells("A3:K3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        if invflag == 0:
            sheet.title = "List of All Cancelled Invoices"
            sheet["A3"] = "List of All Cancelled Invoices"
        elif invflag == 1:
            sheet.title = "List of Cancelled Sales Invoices"
            sheet["A3"] = "List of Cancelled Sales Invoices"
        elif invflag == 2:
            sheet.title = "List of Cancelled Purchase Invoices"
            sheet["A3"] = "List of Cancelled Purchase Invoices"
        sheet.merge_cells("A4:K4")
        sheet["A4"] = (
            "Period: "
            + datetime.strptime(self.request.params["fromdate"], "%Y-%m-%d").strftime(
                "%d-%m-%Y"
            )
            + " to "
            + datetime.strptime(self.request.params["todate"], "%Y-%m-%d").strftime(
                "%d-%m-%Y"
            )
        )

        sheet["A4"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A5"] = "Sr. No."
        sheet["B5"] = "INV No."
        sheet["C5"] = "INV Date"
        sheet["D5"] = "Deli. Note "
        if invflag == 0:
            sheet["E5"] = "Cust/Supp Name"
        elif invflag == 1:
            sheet["E5"] = "Customer Name"
        elif invflag == 2:
            sheet["E5"] = "Supplier Name"
        sheet["F5"] = "Gross Amt"
        sheet["G5"] = "Net Amt"
        sheet["H5"] = "Tax Amt"
        sheet["I5"] = "Godown"
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["A5"].alignment = Alignment(horizontal="center")
        sheet["B5"].alignment = Alignment(horizontal="center")
        sheet["C5"].alignment = Alignment(horizontal="center")
        sheet["D5"].alignment = Alignment(horizontal="center")
        sheet["E5"].alignment = Alignment(horizontal="center")
        sheet["F5"].alignment = Alignment(horizontal="right")
        sheet["G5"].alignment = Alignment(horizontal="right")
        sheet["H5"].alignment = Alignment(horizontal="right")
        sheet["I5"].alignment = Alignment(horizontal="center")
        sheet["A5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["B5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["C5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["D5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["E5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["F5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["G5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["H5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["I5"].font = Font(name="Liberation Serif", size=12, bold=True)
        row = 6
        # Looping each dictionaries in list result to store data in cells and apply styles.
        for invoice in result:
            sheet["A" + str(row)] = invoice["srno"]
            sheet["A" + str(row)].alignment = Alignment(horizontal="center")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["B" + str(row)] = invoice["invoiceno"]
            sheet["B" + str(row)].alignment = Alignment(horizontal="center")
            sheet["B" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["C" + str(row)] = invoice["invoicedate"]
            sheet["C" + str(row)].alignment = Alignment(horizontal="center")
            sheet["C" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            if invoice["dcno"] != "" and invoice["dcdate"] != "":
                sheet["D" + str(row)] = invoice["dcno"] + "," + invoice["dcdate"]
                sheet["D" + str(row)].alignment = Alignment(horizontal="center")
                sheet["D" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            sheet["E" + str(row)] = invoice["custname"]
            sheet["E" + str(row)].alignment = Alignment(horizontal="center")
            sheet["E" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["F" + str(row)] = float("%.2f" % float(invoice["grossamt"]))
            sheet["F" + str(row)].number_format = "0.00"
            sheet["F" + str(row)].alignment = Alignment(horizontal="right")
            sheet["F" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["G" + str(row)] = float("%.2f" % float(invoice["netamt"]))
            sheet["G" + str(row)].number_format = "0.00"
            sheet["G" + str(row)].alignment = Alignment(horizontal="right")
            sheet["G" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["H" + str(row)] = float("%.2f" % float(invoice["taxamt"]))
            sheet["H" + str(row)].number_format = "0.00"
            sheet["H" + str(row)].alignment = Alignment(horizontal="right")
            sheet["H" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["I" + str(row)] = invoice["godown"]
            sheet["I" + str(row)].alignment = Alignment(horizontal="center")
            sheet["I" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            row = row + 1
        output = io.BytesIO()
        invoicewb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true ;path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
        return Response(contents, headerlist=list(headerList.items()))

    except Exception as e:

        print(e)
        return {"gkstatus": 3}


def outstanding_invoices(self):
    """Generate spreadsheet for outstanding invoices

    params
    ======
    orgname: string
    fystart: dd-mm-yyyy
    fyend: dd-mm-yyyy
    fromdate: dd-mm-yyyy
    todate: dd-mm-yyyy
    typeflag: int (valid values: 1, 3, 4)
    inoutflag: int (valid values: 9, 15)
    orderflag: int (valid values: 1, 4)
    """
    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        inoutflag = int(self.request.params["inoutflag"])
        orderflag = int(self.request.params["orderflag"])
        typeflag = int(self.request.params["typeflag"])
        inouts = {9: "Purchase", 15: "Sale"}
        orders = {1: "Ascending", 4: "Descending"}
        types = {1: "Amount Wise", 3: "Party Wise", 4: "Due Wise"}
        inout = inouts[inoutflag]
        order = orders[orderflag]
        reporttype = types[typeflag]
        startdate = self.request.params["fromdate"]
        enddate = self.request.params["todate"]
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        orgname = str(self.request.params["orgname"])
        result = gk_api(
            f"/billwise?type=onlybillsforall&inoutflag={inoutflag}&orderflag={orderflag}&typeflag={typeflag}&startdate={startdate}&enddate={enddate}",
            header,
            self.request,
        )
        # A workbook is opened.
        billwisewb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = billwisewb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "List of Unpaid Invoices"
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 16
        sheet.column_dimensions["F"].width = 16
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells("A1:F2")
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet.merge_cells("A3:F3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        invtype = "Sale"
        if inoutflag == 9:
            invtype = "Purchase"
        sheet["A3"] = "%s List of Outstanding %s Invoices in %s Order" % (
            str(reporttype),
            str(inout),
            str(order),
        )
        sheet.merge_cells("A4:F4")
        sheet["A4"] = (
            "Period: "
            + str(self.request.params["fromdate"])
            + " to "
            + str(self.request.params["todate"])
        )
        sheet["A4"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A5"] = "Sr. No. "
        sheet["B5"] = "Invoice No"
        sheet["C5"] = "Invoice Date"
        custhead = "Customer Name"
        if inoutflag == 9:
            custhead = "Supplier Name"
        sheet["D5"] = custhead
        sheet["E5"] = "Invoice Amount"
        sheet["F5"] = "Amount Pending"
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["E5"].alignment = Alignment(horizontal="right")
        sheet["F5"].alignment = Alignment(horizontal="right")
        sheet["E5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["F5"].font = Font(name="Liberation Serif", size=12, bold=True)
        unAdjInvoices = result["invoices"]
        row = 6
        # Looping each dictionaries in list unAdjInvoices to store data in cells and apply styles.
        srno = 1
        for uninv in unAdjInvoices:
            sheet["A" + str(row)] = srno
            sheet["A" + str(row)].alignment = Alignment(horizontal="left")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["B" + str(row)] = uninv["invoiceno"]
            sheet["B" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["C" + str(row)] = uninv["invoicedate"]
            sheet["C" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["D" + str(row)] = uninv["custname"]
            sheet["D" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["E" + str(row)] = float("%.2f" % float(uninv["invoiceamount"]))
            sheet["E" + str(row)].number_format = "0.00"
            sheet["F" + str(row)] = float("%.2f" % float(uninv["balanceamount"]))
            sheet["F" + str(row)].number_format = "0.00"
            sheet["E" + str(row)].alignment = Alignment(horizontal="right")
            sheet["F" + str(row)].alignment = Alignment(horizontal="right")
            sheet["E" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["F" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            row = row + 1
            srno = srno + 1
        output = io.BytesIO()
        billwisewb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true ;path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print(e)
        return {"gkstatus": 3}
