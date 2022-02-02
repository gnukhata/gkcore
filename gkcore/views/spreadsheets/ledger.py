"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020, 2021 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Survesh" <123survesh@gmail.com>
"Sai Karthik"<kskarthik@disroot.org>

"""
from pyramid.response import Response
from pyramid.request import Request
from datetime import datetime

# Spreadsheet libraries
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import RED

# from io import BytesIO
import io
import json


def print_ledger(self):
    """
    This function returns a spreadsheet form of Ledger Report.
    The spreadsheet in XLSX format is generated by the frontendend.

    params
    ======

    accountcode
    from
    to
    fystart
    fyend
    orgname
    orgtype
    projectcode
    """
    try:
        accountcode = int(self.request.params["accountcode"])
        calculatefrom = self.request.params["from"]
        calculateto = self.request.params["to"]
        fystart = self.request.params["fystart"]
        fyend = self.request.params["fyend"]
        orgname = str(self.request.params["orgname"])
        orgtype = str(self.request.params["orgtype"])
        projectcode = self.request.params["projectcode"]
        header = {"gktoken": self.request.headers["gktoken"]}
        if projectcode == "":
            subreq = Request.blank(
                "/report?type=ledger&accountcode=%d&calculatefrom=%s&calculateto=%s&financialstart=%s&projectcode="
                % (accountcode, calculatefrom, calculateto, fystart),
                headers=header,
            )
            result = self.request.invoke_subrequest(subreq)
        else:
            subreq = Request.blank(
                "/report?type=ledger&accountcode=%d&calculatefrom=%s&calculateto=%s&financialstart=%s&projectcode=%d"
                % (accountcode, calculatefrom, calculateto, fystart, int(projectcode)),
                headers=header,
            )
            result = self.request.invoke_subrequest(subreq)
        headerrow = json.loads(result.text)["ledgerheader"]
        result = json.loads(result.text)["gkresult"]
        ledgerwb = openpyxl.Workbook()
        sheet = ledgerwb.active
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 16
        sheet.column_dimensions["D"].width = 42
        sheet.column_dimensions["E"].width = 16
        sheet.column_dimensions["F"].width = 16
        sheet.column_dimensions["G"].width = 16
        sheet.merge_cells("A1:G2")
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet["A1"] = (
            orgname
            + " (FY: "
            + datetime.strptime(self.request.params["fystart"], "%Y-%m-%d").strftime(
                "%d-%m-%Y"
            )
            + " to "
            + fyend
            + ")"
        )
        sheet.merge_cells("A3:G3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"] = "Account : %s (Period : %s to %s)" % (
            headerrow["accountname"],
            datetime.strptime(self.request.params["from"], "%Y-%m-%d").strftime(
                "%d-%m-%Y"
            ),
            datetime.strptime(self.request.params["to"], "%Y-%m-%d").strftime(
                "%d-%m-%Y"
            ),
        )
        row = 4
        sheet.merge_cells("A4:G4")
        if headerrow["projectname"] != "":
            if orgtype == "Profit Making":
                sheet["A4"].font = Font(name="Liberation Serif", size="14", bold=True)
                sheet["A4"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                sheet["A4"] = "Cost Center : " + headerrow["projectname"]
                row += 1
            else:
                sheet["A4"].font = Font(name="Liberation Serif", size="14", bold=True)
                sheet["A4"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                sheet["A4"] = "Project :" + headerrow["projectname"]
                row += 1
        sheet["A5"] = "Date"
        sheet["B5"] = "V. No."
        sheet["C5"] = "Voucher Type"
        sheet["D5"] = "Particulars"
        sheet["E5"] = "Debit"
        sheet["F5"] = "Credit"
        sheet["G5"] = "Balance"
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["A5"].alignment = Alignment(horizontal="center")
        sheet["B5"].alignment = Alignment(horizontal="center")
        sheet["C5"].alignment = Alignment(horizontal="center")
        sheet["D5"].alignment = Alignment(horizontal="center")
        sheet["E5"].alignment = Alignment(horizontal="right")
        sheet["F5"].alignment = Alignment(horizontal="right")
        sheet["G5"].alignment = Alignment(horizontal="right")
        sheet["A5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["B5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["C5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["D5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["E5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["F5"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["G5"].font = Font(name="Liberation Serif", size=12, bold=True)
        row = 6
        for transaction in result:
            sheet["A" + str(row)] = transaction["voucherdate"]
            sheet["A" + str(row)].alignment = Alignment(horizontal="center")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["B" + str(row)] = transaction["vouchernumber"]
            sheet["B" + str(row)].alignment = Alignment(horizontal="center")
            sheet["B" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )

            if transaction["advflag"] == 1:
                if transaction["Dr"] != "":
                    sheet["E" + str(row)] = float("%.2f" % float(transaction["Dr"]))
                    sheet["E" + str(row)].number_format = "0.00"
                    sheet["E" + str(row)].alignment = Alignment(horizontal="right")
                    # If the advflag = 1 the 'Dr' and 'Cr' amount will be displayed in 'RED' color
                    sheet["E" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=True, color=RED
                    )
                if transaction["Cr"] != "":
                    sheet["F" + str(row)] = float("%.2f" % float(transaction["Cr"]))
                    sheet["F" + str(row)].number_format = "0.00"
                    sheet["F" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["F" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=True, color=RED
                    )
            else:
                if transaction["Dr"] != "":
                    sheet["E" + str(row)] = float("%.2f" % float(transaction["Dr"]))
                    sheet["E" + str(row)].number_format = "0.00"
                    sheet["E" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["E" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                if transaction["Cr"] != "":
                    sheet["F" + str(row)] = float("%.2f" % float(transaction["Cr"]))
                    sheet["F" + str(row)].number_format = "0.00"
                    sheet["F" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["F" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
            sheet["G" + str(row)] = transaction["balance"]
            sheet["G" + str(row)].alignment = Alignment(horizontal="right")
            sheet["G" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )

            if (
                transaction["vouchertype"] == "contra"
                or transaction["vouchertype"] == "purchase"
                or transaction["vouchertype"] == "sales"
                or transaction["vouchertype"] == "receipt"
                or transaction["vouchertype"] == "payment"
                or transaction["vouchertype"] == "journal"
            ):
                sheet["C" + str(row)] = transaction["vouchertype"].title()
                sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                sheet["C" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            elif transaction["vouchertype"] == "debitnote":
                sheet["C" + str(row)] = "Debit Note"
                sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                sheet["C" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            elif transaction["vouchertype"] == "creditnote":
                sheet["C" + str(row)] = "Credit Note"
                sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                sheet["C" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            elif transaction["vouchertype"] == "salesreturn":
                sheet["C" + str(row)] = "Sale Return"
                sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                sheet["C" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            elif transaction["vouchertype"] == "purchasereturn":
                sheet["C" + str(row)] = "Purchase Return"
                sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                sheet["C" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            else:
                sheet["C" + str(row)] = transaction["vouchertype"]
                sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                sheet["C" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            particulars = ""
            length = len(transaction["particulars"])
            for i, k in enumerate(transaction["particulars"]):
                if "amount" in k:
                    sheet["D" + str(row)] = (
                        k["accountname"] + " (" + str(k["amount"]) + ")"
                    )
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                else:
                    sheet["D" + str(row)] = k["accountname"]
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                if i < length - 1:
                    row += 1
            narration = transaction["narration"]
            if narration != "":
                row += 1
                sheet["D" + str(row)] = "(" + narration + ")"
                sheet["D" + str(row)].font = Font(
                    name="Liberation Serif", size="9", italic=True
                )
                sheet["D" + str(row)].alignment = Alignment(vertical="center")
            if "dcinfo" in transaction:
                dcin = transaction["dcinfo"]
                row += 1
                sheet["D" + str(row)] = "(" + dcin + ")"
                sheet["D" + str(row)].font = Font(
                    name="Liberation Serif", size="9", italic=True
                )
                sheet["D" + str(row)].alignment = Alignment(vertical="center")
            row += 1
        output = io.BytesIO()
        ledgerwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
        }
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print(e)
        return {"gkstatus": 3}


def print_monthly_ledger(self):

    """
    This function returns a spreadsheet form of Monthly Ledger Report.
    The spreadsheet in XLSX format is generated by the frontendend.

    params:
    =======
    accountccode
    fystart
    fyend
    accname
    fyend
    orgname

    """
    try:
        accountcode = int(self.request.params["accountcode"])
        fystart = str(self.request.params["fystart"])
        accountname = str(self.request.params["accname"])
        fyend = self.request.params["fyend"]
        orgname = str(self.request.params["orgname"])
        header = {"gktoken": self.request.headers["gktoken"]}
        subreq = Request.blank(
            "/report?type=monthlyledger&accountcode=%d" % (accountcode),
            headers=header,
        )
        # result = result.json()["gkresult"]
        result = json.loads(self.request.invoke_subrequest(subreq).text)["gkresult"]
        fystart = datetime.strptime(
            self.request.params["fystart"], "%Y-%m-%d"
        ).strftime("%d-%m-%Y")
        mledgerwb = openpyxl.Workbook()
        sheet = mledgerwb.active
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 18
        sheet.column_dimensions["D"].width = 18
        sheet.column_dimensions["E"].width = 20
        sheet.column_dimensions["F"].width = 24
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells("A1:F2")
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet.merge_cells("A3:F3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"] = "Account : %s" % (accountname)
        sheet["A4"] = "Month"
        sheet["B4"] = "Debit Balance"
        sheet["C4"] = "No. of Debit Records"
        sheet["D4"] = "Credit Balance"
        sheet["E4"] = "No. of Credit Records"
        sheet["F4"] = "No. of Unlocked Transactions"
        titlerow = sheet.row_dimensions[4]
        titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["A4"].alignment = Alignment(horizontal="center")
        sheet["B4"].alignment = Alignment(horizontal="right")
        sheet["C4"].alignment = Alignment(horizontal="center")
        sheet["D4"].alignment = Alignment(horizontal="right")
        sheet["E4"].alignment = Alignment(horizontal="center")
        sheet["F4"].alignment = Alignment(horizontal="center")
        sheet["A4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["B4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["C4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["D4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["E4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["F4"].font = Font(name="Liberation Serif", size=12, bold=True)
        row = 5
        # Looping each dictionaries in list result to store data in cells and apply styles.
        for eachmonth in result:
            sheet["A" + str(row)] = eachmonth["month"]
            sheet["A" + str(row)].alignment = Alignment(horizontal="center")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            if eachmonth["advflag"] == 1:
                if eachmonth["Dr"] != "":
                    sheet["B" + str(row)] = float("%.2f" % float(eachmonth["Dr"]))
                    sheet["B" + str(row)].number_format = "0.00"
                    sheet["B" + str(row)].alignment = Alignment(horizontal="right")
                    # If the advflag = 1 the 'Dr' and 'Cr' amount will be displayed in 'RED' color
                    sheet["B" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=True, color=RED
                    )
                sheet["C" + str(row)] = eachmonth["vcountDr"]
                sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                sheet["C" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["D" + str(row)] = eachmonth["Cr"]
                if eachmonth["Cr"] != "":
                    sheet["D" + str(row)] = float("%.2f" % float(eachmonth["Cr"]))
                    sheet["D" + str(row)].number_format = "0.00"
                    sheet["D" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=True, color=RED
                    )
                sheet["E" + str(row)] = eachmonth["vcountCr"]
                sheet["E" + str(row)].alignment = Alignment(horizontal="center")
                sheet["E" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["F" + str(row)] = eachmonth["vcount"] - eachmonth["vcountLock"]
                sheet["F" + str(row)].alignment = Alignment(horizontal="center")
                sheet["F" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            else:
                if eachmonth["Dr"] != "":
                    sheet["B" + str(row)] = float("%.2f" % float(eachmonth["Dr"]))
                    sheet["B" + str(row)].number_format = "0.00"
                    sheet["B" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["B" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                sheet["C" + str(row)] = eachmonth["vcountDr"]
                sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                sheet["C" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["D" + str(row)] = eachmonth["Cr"]
                if eachmonth["Cr"] != "":
                    sheet["D" + str(row)] = float("%.2f" % float(eachmonth["Cr"]))
                    sheet["D" + str(row)].number_format = "0.00"
                    sheet["D" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                sheet["E" + str(row)] = eachmonth["vcountCr"]
                sheet["E" + str(row)].alignment = Alignment(horizontal="center")
                sheet["E" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["F" + str(row)] = eachmonth["vcount"] - eachmonth["vcountLock"]
                sheet["F" + str(row)].alignment = Alignment(horizontal="center")
                sheet["F" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            row += 1
        output = io.BytesIO()
        mledgerwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true ;path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except:
        print("File not found")
        return {"gkstatus": 3}
