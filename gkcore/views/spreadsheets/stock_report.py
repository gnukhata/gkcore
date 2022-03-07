"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Survesh" <123survesh@gmail.com>
"Sai Karthik"<kskarthik@disroot.org>

"""

# from gkcore import eng, enumdict
# from gkcore.views.api_login import authCheck
# from gkcore.models.gkdb import (
#     organisation,
#     unitofmeasurement,
#     product,
#     goprod,
#     stock,
#     categorysubcategories,
# )
# from sqlalchemy.sql import select
# from sqlalchemy.engine.base import Connection
# from sqlalchemy import and_, exc, func
from pyramid.request import Request
from pyramid.response import Response

# from pyramid.view import view_defaults, view_config
# import gkcore
# from gkcore.views.api_reports import getBalanceSheet
# from gkcore.views.api_invoice import getInvoiceList
from datetime import datetime

# from gkcore.views.api_user import getUserRole
# from gkcore.views.api_godown import getusergodowns
# import requests
import json

# Spreadsheet libraries
import openpyxl
from openpyxl.styles import Font, Alignment

# from openpyxl.styles.colors import RED


# from io import BytesIO
import io


def print_stock_report(self):
    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        godownflag = int(self.request.params["godownflag"])
        if godownflag == 1:
            goaddr = self.request.params["goaddr"]
            goid = int(self.request.params["goid"])
            goname = self.request.params["goname"]
        productcode = int(self.request.params["productcode"])
        calculatefrom = self.request.params["calculatefrom"]
        calculateto = self.request.params["calculateto"]
        scalculatefrom = datetime.strptime(calculatefrom, "%d-%m-%Y").strftime(
            "%Y-%m-%d"
        )
        scalculateto = datetime.strptime(calculateto, "%d-%m-%Y").strftime("%Y-%m-%d")
        productdesc = self.request.params["productdesc"]
        if godownflag > 0:
            subreq = Request.blank(
                "/report?type=godownstockreport&productcode=%d&startdate=%s&enddate=%s&goid=%d&godownflag=%d"
                % (productcode, scalculatefrom, scalculateto, goid, godownflag),
                headers=header,
            )
            result = self.request.invoke_subrequest(subreq)
        else:
            subr = Request.blank(
                "/report?type=stockreport&productcode=%d&startdate=%s&enddate=%s"
                % (productcode, scalculatefrom, scalculateto),
                headers=header,
            )
            result = self.request.invoke_subrequest(subr)
        result = json.loads(result.text)["gkresult"]
        fystart = self.request.params["fystart"]
        # fystart = datetime.strptime(
        #     self.request.params["fystart"], "%Y-%m-%d"
        # ).strftime("%d-%m-%Y")
        fyend = str(self.request.params["fyend"])
        orgname = str(self.request.params["orgname"])
        # A workbook is opened.
        prowb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = prowb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "Product Report"
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 22
        sheet.column_dimensions["D"].width = 18
        sheet.column_dimensions["E"].width = 14
        sheet.column_dimensions["F"].width = 14
        sheet.column_dimensions["G"].width = 14
        sheet.column_dimensions["H"].width = 14
        sheet.column_dimensions["I"].width = 14
        sheet.column_dimensions["J"].width = 14
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells("A1:I2")
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet.merge_cells("A3:J3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells("A4:J4")
        sheet["A4"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
        trn = ""
        if godownflag > 0:
            sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
            sheet["A3"] = (
                "Godown Wise Product Report  (Period : "
                + calculatefrom
                + " to "
                + calculateto
                + ")"
            )
            sheet["A4"] = "Name of the Product: " + productdesc
            sheet.merge_cells("A5:J5")
            sheet["A5"].font = Font(name="Liberation Serif", size="14", bold=True)
            sheet["A5"] = (
                "Name of the Godown : " + goname + ", Godown Address: " + goaddr
            )
            sheet["A6"] = "Date"
            sheet["B6"] = "Particulars"
            sheet["C6"] = "Document Type"
            sheet["D6"] = "Deli Note No."
            sheet["E6"] = "INV/DR/CR No."
            sheet["F6"] = "RN No."
            sheet["G6"] = "TN No."
            sheet["H6"] = "Inward"
            sheet["I6"] = "Outward"
            sheet["J6"] = "Balance"
            titlerow = sheet.row_dimensions[6]
            titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
            titlerow.alignment = Alignment(horizontal="center", vertical="center")
            sheet["H6"].alignment = Alignment(horizontal="right")
            sheet["H6"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["I6"].alignment = Alignment(horizontal="right")
            sheet["I6"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["J6"].alignment = Alignment(horizontal="right")
            sheet["J6"].font = Font(name="Liberation Serif", size=12, bold=True)
            row = 7

            for stock in result:
                if stock["trntype"] == "delchal":
                    trn = "Delivery Note"
                if stock["trntype"] == "invoice":
                    trn = "Invoice "
                if stock["trntype"] == "delchal&invoice":
                    trn = "Delivery Note & Invoice"
                if stock["trntype"] == "transfer note":
                    trn = "Transfer Note "
                if stock["trntype"] == "Rejection Note":
                    trn = "Rejection Note"
                if stock["trntype"] == "Debit Note":
                    trn = "Debit Note"
                if stock["trntype"] == "Credit Note":
                    trn = "Credit Note"

                if (
                    stock["particulars"] == "opening stock"
                    and stock["dcno"] == ""
                    and stock["invno"] == ""
                    and stock["date"] == ""
                ):
                    sheet["A" + str(row)] = ""
                    sheet["A" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["A" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["B" + str(row)] = stock["particulars"].title()
                    sheet["B" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)] = ""
                    sheet["C" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["D" + str(row)] = ""
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["D" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["E" + str(row)] = ""
                    sheet["E" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["E" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["F" + str(row)] = ""
                    sheet["F" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["F" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["G" + str(row)] = ""
                    sheet["G" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["G" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["H" + str(row)] = float("%.2f" % float(stock["inward"]))
                    sheet["H" + str(row)].number_format = "0.00"
                    sheet["H" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["H" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["I" + str(row)] = ""
                    sheet["I" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["I" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["J" + str(row)] = ""
                    sheet["J" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["J" + str(row)].alignment = Alignment(horizontal="right")
                if (
                    stock["particulars"] != "Total"
                    and (
                        stock["dcno"] != ""
                        or stock["invno"] != ""
                        or stock["tnno"] != ""
                        or stock["rnno"] != ""
                    )
                    and stock["date"] != ""
                ):

                    sheet["A" + str(row)] = stock["date"]
                    sheet["A" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["A" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["B" + str(row)] = stock["particulars"]
                    sheet["B" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)] = trn
                    sheet["C" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["D" + str(row)] = stock["dcno"]
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["D" + str(row)].alignment = Alignment(horizontal="center")
                    if stock["invno"] != "":

                        sheet["E" + str(row)] = stock["invno"]
                    elif "drcrno" in stock and stock["drcrno"] != "":

                        sheet["E" + str(row)] = stock["drcrno"]
                    else:

                        sheet["E" + str(row)] = ""
                    sheet["E" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["E" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["F" + str(row)] = stock["rnno"]
                    sheet["F" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["F" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["G" + str(row)] = stock["tnno"]
                    sheet["G" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["G" + str(row)].alignment = Alignment(horizontal="center")
                    if stock["inwardqty"] != "":

                        sheet["H" + str(row)] = float(
                            "%.2f" % float(stock["inwardqty"])
                        )
                        sheet["H" + str(row)].number_format = "0.00"
                        sheet["H" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        sheet["H" + str(row)].alignment = Alignment(horizontal="right")
                    if stock["outwardqty"] != "":

                        sheet["I" + str(row)] = float(
                            "%.2f" % float(stock["outwardqty"])
                        )
                        sheet["I" + str(row)].number_format = "0.00"
                        sheet["I" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        sheet["I" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["J" + str(row)] = float("%.2f" % float(stock["balance"]))
                    sheet["J" + str(row)].number_format = "0.00"
                    sheet["J" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["J" + str(row)].alignment = Alignment(horizontal="right")
                if (
                    stock["particulars"] == "Total"
                    and stock["dcno"] == ""
                    and stock["invno"] == ""
                    and stock["date"] == ""
                ):

                    sheet["A" + str(row)] = ""
                    sheet["A" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["A" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["B" + str(row)] = stock["particulars"].title()
                    sheet["B" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)] = ""
                    sheet["C" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["D" + str(row)] = ""
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["D" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["E" + str(row)] = ""
                    sheet["E" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["E" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["F" + str(row)] = ""
                    sheet["F" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["F" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["G" + str(row)] = ""
                    sheet["G" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["G" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["H" + str(row)] = float(
                        "%.2f" % float(stock["totalinwardqty"])
                    )
                    sheet["H" + str(row)].number_format = "0.00"
                    sheet["H" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["H" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["I" + str(row)] = float(
                        "%.2f" % float(stock["totaloutwardqty"])
                    )
                    sheet["I" + str(row)].number_format = "0.00"
                    sheet["I" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["I" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["J" + str(row)] = ""
                    sheet["J" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["J" + str(row)].alignment = Alignment(horizontal="right")
                row += 1
        else:
            sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
            sheet["A3"] = (
                "Product Report (Period : " + calculatefrom + " to " + calculateto + ")"
            )
            sheet["A4"] = "Name of the Product: " + productdesc
            sheet["A5"] = "Date"
            sheet["B5"] = "Particulars"
            sheet["C5"] = "Document Type"
            sheet["D5"] = "Deli Note No."
            sheet["E5"] = "INV/DR/CR No."
            sheet["F5"] = "RN No."
            sheet["G5"] = "Inward"
            sheet["H5"] = "Outward"
            sheet["I5"] = "Balance"
            titlerow = sheet.row_dimensions[5]
            titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
            titlerow.alignment = Alignment(horizontal="center", vertical="center")
            sheet["G5"].alignment = Alignment(horizontal="right")
            sheet["G5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["H5"].alignment = Alignment(horizontal="right")
            sheet["H5"].font = Font(name="Liberation Serif", size=12, bold=True)
            sheet["I5"].alignment = Alignment(horizontal="right")
            sheet["I5"].font = Font(name="Liberation Serif", size=12, bold=True)
            row = 6

            for stock in result:
                if stock["trntype"] == "delchal":
                    trn = "Delivery Note"
                if stock["trntype"] == "invoice":
                    trn = "Invoice "
                if stock["trntype"] == "delchal&invoice":
                    trn = "Delivery Note & Invoice"
                if stock["trntype"] == "transfer note":
                    trn = "Transfer Note "
                if stock["trntype"] == "Rejection Note":
                    trn = "Rejection Note"
                if stock["trntype"] == "Debit Note":
                    trn = "Debit Note"
                if stock["trntype"] == "Credit Note":
                    trn = "Credit Note"

                if (
                    stock["particulars"] == "opening stock"
                    and stock["dcno"] == ""
                    and stock["invno"] == ""
                    and stock["date"] == ""
                ):
                    sheet["A" + str(row)] = ""
                    sheet["A" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["A" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["B" + str(row)] = stock["particulars"].title()
                    sheet["B" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)] = ""
                    sheet["C" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["D" + str(row)] = ""
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["D" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["E" + str(row)] = ""
                    sheet["E" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["E" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["F" + str(row)] = ""
                    sheet["F" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["F" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["G" + str(row)] = float("%.2f" % float(stock["inward"]))
                    sheet["G" + str(row)].number_format = "0.00"
                    sheet["G" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["G" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["H" + str(row)] = ""
                    sheet["H" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["H" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["I" + str(row)] = ""
                    sheet["I" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["I" + str(row)].alignment = Alignment(horizontal="right")
                if (
                    stock["particulars"] != "Total"
                    and (
                        stock["dcno"] != ""
                        or stock["invno"] != ""
                        or stock["rnid"] != ""
                        or stock["drcrno"] != ""
                    )
                    and stock["date"] != ""
                ):
                    sheet["A" + str(row)] = stock["date"]
                    sheet["A" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["A" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["B" + str(row)] = stock["particulars"]
                    sheet["B" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )

                    sheet["C" + str(row)] = trn
                    sheet["C" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )

                    sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["D" + str(row)] = stock["dcno"]
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["D" + str(row)].alignment = Alignment(horizontal="center")
                    if stock["invno"] != "":
                        sheet["E" + str(row)] = stock["invno"]
                    elif "drcrno" in stock and stock["drcrno"] != "":
                        sheet["E" + str(row)] = stock["drcrno"]
                    else:
                        sheet["E" + str(row)] = ""
                    sheet["E" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["E" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["F" + str(row)] = stock["rnno"]
                    sheet["F" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["F" + str(row)].alignment = Alignment(horizontal="right")
                    if stock["inwardqty"] != "":
                        sheet["G" + str(row)] = float(
                            "%.2f" % float(stock["inwardqty"])
                        )
                        sheet["G" + str(row)].number_format = "0.00"
                        sheet["G" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        sheet["G" + str(row)].alignment = Alignment(horizontal="right")
                    if stock["outwardqty"] != "":
                        sheet["H" + str(row)] = float(
                            "%.2f" % float(stock["outwardqty"])
                        )
                        sheet["H" + str(row)].number_format = "0.00"
                        sheet["H" + str(row)].font = Font(
                            name="Liberation Serif", size="12", bold=False
                        )
                        sheet["H" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["I" + str(row)] = float("%.2f" % float(stock["balance"]))
                    sheet["I" + str(row)].number_format = "0.00"
                    sheet["I" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["I" + str(row)].alignment = Alignment(horizontal="right")
                if (
                    stock["particulars"] == "Total"
                    and stock["dcno"] == ""
                    and stock["invno"] == ""
                    and stock["date"] == ""
                ):
                    sheet["A" + str(row)] = ""
                    sheet["A" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["A" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["B" + str(row)] = stock["particulars"].title()
                    sheet["B" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)] = ""
                    sheet["C" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["C" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["D" + str(row)] = ""
                    sheet["D" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["D" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["E" + str(row)] = ""
                    sheet["E" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["E" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["F" + str(row)] = ""
                    sheet["F" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["F" + str(row)].alignment = Alignment(horizontal="center")
                    sheet["G" + str(row)] = float(
                        "%.2f" % float(stock["totalinwardqty"])
                    )
                    sheet["G" + str(row)].number_format = "0.00"
                    sheet["G" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["G" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["H" + str(row)] = float(
                        "%.2f" % float(stock["totaloutwardqty"])
                    )
                    sheet["H" + str(row)].number_format = "0.00"
                    sheet["H" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["H" + str(row)].alignment = Alignment(horizontal="right")
                    sheet["I" + str(row)] = ""
                    sheet["I" + str(row)].font = Font(
                        name="Liberation Serif", size="12", bold=False
                    )
                    sheet["I" + str(row)].alignment = Alignment(horizontal="right")
                row += 1
        output = io.BytesIO()
        prowb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}

        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print("exception:", e)
        return {"gkstatus": 3}
