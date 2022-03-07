"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020, 2021 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors
============
Survesh VRL <123survesh@gmail.com>
Sai Karthik <kskarthik@disroot.org>

"""

from gkcore.models.meta import gk_api
from pyramid.response import Response

# Spreadsheet libraries
import io
import openpyxl
from openpyxl.styles import Font, Alignment


def cash_report(self):
    """Generate spreadsheet for budget cash report

    params
    ======
    orgname: string
    fystart: dd-mm-yyyy
    fyend: dd-mm-yyyy
    financialstart: yyyy-mm-dd
    budid = budget id
    budgetdetails = string
    """

    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        financialstart = self.request.params["financialstart"]
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        orgname = str(self.request.params["orgname"])
        budgetdetails = str(self.request.params["budgetdetails"])
        result = gk_api(
            "/budget?type=cashReport&budid=%d&financialstart=%s"
            % (int(self.request.params["budid"]), str(financialstart)),
            header,
            self.request,
        )['gkresult']
        budgetwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = budgetwb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "Cash Budget Report"

        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 20
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells("A1:E2")
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"] = "Cash Budget Report : " + str(budgetdetails)
        sheet.merge_cells("A3:E3")

        sheet["A4"].font = Font(name="Liberation Serif", size="12", bold=True)
        sheet["A4"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A4"] = ""
        sheet.merge_cells("A4:E4")

        sheet["A5"] = "Particulars"
        sheet["B5"] = "Budgeted"
        sheet["C5"] = "Actuals"
        sheet["D5"] = "Variance"
        sheet["E5"] = "Variance (%)"
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name="Liberation Serif", size="12", bold=True)
        titlerow.alignment = Alignment(horizontal="center", vertical="center")

        titlecolumn = sheet.column_dimensions["A"]
        titlecolumn.font = Font(name="Liberation Serif", size="12", bold=True)
        titlecolumn.alignment = Alignment(horizontal="center", vertical="center")
        sheet["A6"] = "Opening"
        row = 7
        for ob in result["openingacc"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name="Liberation Serif")
            accountsrow.alignment = Alignment(horizontal="right", vertical="center")
            sheet["A" + str(row)] = ob["accountname"]
            sheet["A" + str(row)].font = Font(name="Liberation Serif", italic=True)
            sheet["A" + str(row)].alignment = Alignment(
                horizontal="right", vertical="center"
            )
            sheet["B" + str(row)] = float(ob["balance"])
            sheet["B" + str(row)].number_format = "0.00"
            sheet["C" + str(row)] = float(ob["balance"])
            sheet["C" + str(row)].number_format = "0.00"
            sheet["D" + str(row)] = "-"
            sheet["E" + str(row)] = "-"
            row = row + 1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name="Liberation Serif", bold=True)
        totalrow.alignment = Alignment(horizontal="right", vertical="center")
        sheet["A" + str(row)] = "Total"
        sheet["A" + str(row)].alignment = Alignment(
            horizontal="right", vertical="center"
        )
        sheet["A" + str(row)].font = Font(name="Liberation Serif", bold=True)
        sheet["B" + str(row)] = float(result["opening"])
        sheet["B" + str(row)].number_format = "0.00"
        sheet["C" + str(row)] = float(result["opening"])
        sheet["C" + str(row)].number_format = "0.00"
        sheet["D" + str(row)] = "-"
        sheet["E" + str(row)] = "-"
        row = row + 1
        sheet["A" + str(row)] = "Inflow"
        row = row + 1
        for ob in result["inflow"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name="Liberation Serif")
            accountsrow.alignment = Alignment(horizontal="right", vertical="center")
            sheet["A" + str(row)] = ob["accountname"]
            sheet["A" + str(row)].font = Font(name="Liberation Serif", italic=True)
            sheet["A" + str(row)].alignment = Alignment(
                horizontal="right", vertical="center"
            )
            sheet["B" + str(row)] = float(ob["budget"])
            sheet["B" + str(row)].number_format = "0.00"
            sheet["C" + str(row)] = float(ob["actual"])
            sheet["C" + str(row)].number_format = "0.00"
            if ob["var"] != "-":
                sheet["D" + str(row)] = float(ob["var"])
                sheet["D" + str(row)].number_format = "0.00"
            else:
                sheet["D" + str(row)] = ob["var"]
            if ob["varinpercent"] == "-":
                sheet["E" + str(row)] = ob["varinpercent"]
            else:
                sheet["E" + str(row)] = ob["varinpercent"] + " %"
            row = row + 1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name="Liberation Serif", bold=True)
        totalrow.alignment = Alignment(horizontal="right", vertical="center")
        sheet["A" + str(row)] = "Total"
        sheet["A" + str(row)].alignment = Alignment(
            horizontal="right", vertical="center"
        )
        sheet["A" + str(row)].font = Font(name="Liberation Serif", bold=True)
        sheet["B" + str(row)] = float(result["budgetin"])
        sheet["B" + str(row)].number_format = "0.00"
        sheet["C" + str(row)] = float(result["actualin"])
        sheet["C" + str(row)].number_format = "0.00"
        sheet["D" + str(row)] = float(result["varin"])
        sheet["D" + str(row)].number_format = "0.00"
        sheet["E" + str(row)] = result["varpercentin"] + " %"
        row = row + 1
        sheet["A" + str(row)] = "Outflow"
        row = row + 1
        for ob in result["outflow"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name="Liberation Serif")
            accountsrow.alignment = Alignment(horizontal="right", vertical="center")
            sheet["A" + str(row)] = ob["accountname"]
            sheet["A" + str(row)].font = Font(name="Liberation Serif", italic=True)
            sheet["A" + str(row)].alignment = Alignment(
                horizontal="right", vertical="center"
            )
            sheet["B" + str(row)] = float(ob["budget"])
            sheet["B" + str(row)].number_format = "0.00"
            sheet["C" + str(row)] = float(ob["actual"])
            sheet["C" + str(row)].number_format = "0.00"
            if ob["var"] != "-":
                sheet["D" + str(row)] = float(ob["var"])
                sheet["D" + str(row)].number_format = "0.00"
            else:
                sheet["D" + str(row)] = ob["var"]
            if ob["varinpercent"] == "-":
                sheet["E" + str(row)] = ob["varinpercent"]
            else:
                sheet["E" + str(row)] = ob["varinpercent"] + " %"
            row = row + 1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name="Liberation Serif", bold=True)
        totalrow.alignment = Alignment(horizontal="right", vertical="center")
        sheet["A" + str(row)] = "Total"
        sheet["A" + str(row)].alignment = Alignment(
            horizontal="right", vertical="center"
        )
        sheet["A" + str(row)].font = Font(name="Liberation Serif", bold=True)
        sheet["B" + str(row)] = float(result["budgetout"])
        sheet["B" + str(row)].number_format = "0.00"
        sheet["C" + str(row)] = float(result["actualout"])
        sheet["C" + str(row)].number_format = "0.00"
        sheet["D" + str(row)] = float(result["varout"])
        sheet["D" + str(row)].number_format = "0.00"
        sheet["E" + str(row)] = result["varpercentout"] + " %"
        row = row + 1
        sheet["A" + str(row)] = "Closing"
        row = row + 1
        for ob in result["closing"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name="Liberation Serif")
            accountsrow.alignment = Alignment(horizontal="right", vertical="center")
            sheet["A" + str(row)] = ob["accountname"]
            sheet["A" + str(row)].font = Font(name="Liberation Serif", italic=True)
            sheet["A" + str(row)].alignment = Alignment(
                horizontal="right", vertical="center"
            )
            sheet["B" + str(row)] = float(ob["budget"])
            sheet["B" + str(row)].number_format = "0.00"
            sheet["C" + str(row)] = float(ob["balance"])
            sheet["C" + str(row)].number_format = "0.00"
            if ob["var"] != "-":
                sheet["D" + str(row)] = float(ob["var"])
                sheet["D" + str(row)].number_format = "0.00"
            else:
                sheet["D" + str(row)] = ob["var"]
            if ob["varinpercent"] == "-":
                sheet["E" + str(row)] = ob["varinpercent"]
            else:
                sheet["E" + str(row)] = ob["varinpercent"] + " %"
            row = row + 1

        output = io.BytesIO()
        budgetwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true ;path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print(e)
        return {"gkstatus": 3}


def pnl(self):
    """Generate spreadsheet for budget P&L

    params
    ======
    orgname: string
    fystart: dd-mm-yyyy
    fyend: dd-mm-yyyy
    financialstart: yyyy-mm-dd
    budid = budget id
    budgetdetails = string
    """

    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        financialstart = self.request.params["financialstart"]
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        Otype = str(self.request.params["orgtype"])
        orgname = str(self.request.params["orgname"])
        budgetdetails = str(self.request.params["budgetdetails"])
        result = gk_api(
            "/budget?type=profitlossReport&budid=%d&financialstart=%s"
            % (int(self.request.params["budid"]), str(financialstart)),
            header,
            self.request,
        )["gkresult"]
        budgetwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = budgetwb.active
        # Title of the sheet and width of columns are set.
        if Otype == "Not For Profit":
            sheet.title = "Income and Expenditure Budget Report"
        else:
            sheet.title = "Profit & Loss Budget Report"

        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 20
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells("A1:E2")
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        if Otype == "Not For Profit":
            sheet["A3"] = "Income and Expenditure Budget Report :" + str(budgetdetails)
        else:
            sheet["A3"] = "Profit & Loss Budget Report :" + str(budgetdetails)

        sheet.merge_cells("A3:E3")

        sheet["A4"].font = Font(name="Liberation Serif", size="12", bold=True)
        sheet["A4"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A4"] = ""
        sheet.merge_cells("A4:E4")
        sheet["A5"] = "Particulars"
        sheet["B5"] = "Budgeted"
        sheet["C5"] = "Actuals"
        sheet["D5"] = "Variance"
        sheet["E5"] = "Variance (%)"
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name="Liberation Serif", size="12", bold=True)
        titlerow.alignment = Alignment(horizontal="center", vertical="center")
        sheet["A6"] = "Incomes"
        sheet["A6"].font = Font(name="Liberation Serif", bold=True)
        sheet["A6"].alignment = Alignment(horizontal="left", vertical="center")
        row = 7
        for budget in result["incomeacc"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name="Liberation Serif")
            accountsrow.alignment = Alignment(horizontal="right", vertical="center")
            sheet["A" + str(row)] = budget["name"]
            sheet["A" + str(row)].font = Font(italic=True, name="Liberation Serif")
            sheet["A" + str(row)].alignment = Alignment(
                horizontal="right", vertical="center"
            )
            sheet["B" + str(row)] = float(budget["budget"])
            sheet["B" + str(row)].number_format = "0.00"
            sheet["C" + str(row)] = float(budget["actual"])
            sheet["C" + str(row)].number_format = "0.00"
            if budget["var"] != "-":
                sheet["D" + str(row)] = float(budget["var"])
                sheet["D" + str(row)].number_format = "0.00"
            else:
                sheet["D" + str(row)] = budget["var"]
            if budget["varinpercent"] == "-":
                sheet["E" + str(row)] = budget["varinpercent"]
            else:
                sheet["E" + str(row)] = budget["varinpercent"] + " %"
            row = row + 1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name="Liberation Serif", bold=True)
        totalrow.alignment = Alignment(horizontal="right", vertical="center")
        sheet["A" + str(row)] = "Total "
        sheet["A" + str(row)].font = Font(name="Liberation Serif", bold=True)
        sheet["A" + str(row)].alignment = Alignment(
            horizontal="right", vertical="center"
        )
        sheet["B" + str(row)] = float(result["budgetincome"])
        sheet["B" + str(row)].number_format = "0.00"
        sheet["C" + str(row)] = float(result["income"])
        sheet["C" + str(row)].number_format = "0.00"
        sheet["D" + str(row)] = float(result["varincome"])
        sheet["D" + str(row)].number_format = "0.00"
        sheet["E" + str(row)] = result["varinpercentincome"] + " %"
        row = row + 1
        sheet["A" + str(row)] = "Expenses"
        sheet["A" + str(row)].font = Font(name="Liberation Serif", bold=True)
        sheet["A" + str(row)].alignment = Alignment(
            horizontal="left", vertical="center"
        )
        row = row + 1
        for budget in result["expenseacc"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name="Liberation Serif")
            accountsrow.alignment = Alignment(horizontal="right", vertical="center")
            sheet["A" + str(row)] = budget["name"]
            sheet["A" + str(row)].font = Font(italic=True, name="Liberation Serif")
            sheet["A" + str(row)].alignment = Alignment(
                horizontal="right", vertical="center"
            )
            sheet["B" + str(row)] = float(budget["budget"])
            sheet["B" + str(row)].number_format = "0.00"
            sheet["C" + str(row)] = float(budget["actual"])
            sheet["C" + str(row)].number_format = "0.00"
            if budget["var"] != "-":
                sheet["D" + str(row)] = float(budget["var"])
                sheet["D" + str(row)].number_format = "0.00"
            else:
                sheet["D" + str(row)] = budget["var"]
            if budget["varinpercent"] == "-":
                sheet["E" + str(row)] = budget["varinpercent"]
            else:
                sheet["E" + str(row)] = budget["varinpercent"] + " %"
            row = row + 1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name="Liberation Serif", bold=True)
        totalrow.alignment = Alignment(horizontal="right", vertical="center")
        sheet["A" + str(row)] = "Total "
        sheet["A" + str(row)].font = Font(name="Liberation Serif", bold=True)
        sheet["A" + str(row)].alignment = Alignment(
            horizontal="right", vertical="center"
        )
        sheet["B" + str(row)] = float(result["budgetexpense"])
        sheet["B" + str(row)].number_format = "0.00"
        sheet["C" + str(row)] = float(result["expense"])
        sheet["C" + str(row)].number_format = "0.00"
        sheet["D" + str(row)] = float(result["varexpense"])
        sheet["D" + str(row)].number_format = "0.00"
        sheet["E" + str(row)] = result["varinpercentexp"] + " %"
        row = row + 1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name="Liberation Serif", bold=True)
        totalrow.alignment = Alignment(horizontal="right", vertical="center")
        sheet["A" + str(row)] = "Net Profit"
        sheet["A" + str(row)].font = Font(name="Liberation Serif", bold=True)
        sheet["A" + str(row)].alignment = Alignment(
            horizontal="left", vertical="center"
        )
        sheet["B" + str(row)] = float(result["budgetprofit"])
        sheet["B" + str(row)].number_format = "0.00"
        sheet["C" + str(row)] = float(result["profit"])
        sheet["C" + str(row)].number_format = "0.00"
        if result["varprofit"] != "-":
            sheet["D" + str(row)] = float(result["varprofit"])
            sheet["D" + str(row)].number_format = "0.00"
        else:
            sheet["D" + str(row)] = result["varprofit"]
            if result["varinpercentprofit"] != "-":
                sheet["E" + str(row)] = result["varinpercentprofit"] + " %"
            else:
                sheet["E" + str(row)] = "-"

        output = io.BytesIO()
        budgetwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true ;path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print(e)
        return {"gkstatus": 3}
