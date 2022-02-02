"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Survesh" <123survesh@gmail.com>
"Sai Karthik"<kskarthik@disroot.org>

"""
from gkcore.models.meta import gk_api
from pyramid.response import Response

# Spreadsheet libraries
import openpyxl
from openpyxl.styles import Font, Alignment

# from io import BytesIO
import io


def user_list(self):
    """Return spreadsheet with list of users

    params
    ======
    fystart
    fyend
    orgname

    """
    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        result = gk_api("/users?type=list", header, self.request)
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        orgname = str(self.request.params["orgname"])
        # A workbook is opened.
        userwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = userwb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "List of Users"
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 16
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 40
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells("A1:F2")
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet.merge_cells("A3:F3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"] = "List of Users"
        sheet["A4"] = "Sr. No. "
        sheet["B4"] = "User Name"
        sheet["C4"] = "User Role"
        sheet["D4"] = "Associated Godowns(s)"
        titlerow = sheet.row_dimensions[4]
        titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
        userList = result["gkresult"]
        row = 5
        # Looping to store the data in the cells and apply styles.
        srno = 1
        for user in userList:
            sheet["A" + str(row)] = srno
            sheet["A" + str(row)].alignment = Alignment(horizontal="left")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["B" + str(row)] = user["username"]
            sheet["B" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["C" + str(row)] = user["userrole"]
            sheet["C" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            gostring = ""
            i = 1
            for godown in user["godowns"]:
                if i == user["noofgodowns"]:
                    gostring += godown
                else:
                    gostring = gostring + godown + ","
                i += 1
            sheet["D" + str(row)] = gostring
            sheet["D" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            row = row + 1
            srno = srno + 1
        output = io.BytesIO()
        userwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ ;HttpOnly",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/ ;'}

        return Response(contents, headerlist=list(headerList.items()))
    except:
        return {"gkstatus": 3}
