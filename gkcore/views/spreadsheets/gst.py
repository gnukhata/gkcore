"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020, 2021 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors
============
Survesh VRL <123survesh@gmail.com>
Sai Karthik <kskarthik@disroot.org>

"""

from pyramid.response import Response

# Spreadsheet libraries
import io
import openpyxl
from openpyxl.styles import Font, Alignment

from gkcore.models.meta import gk_api
from gkcore.views.spreadsheets.gst_r1_template import gst_r1_template


def r1(self):
    """Creates the spreadsheet for GSTR-1
    params:
    start = yyyy-mm-dd
    end = yyyy-mm-dd
    """

    header = {"gktoken": self.request.headers["gktoken"]}

    result = gk_api(
        f"/gstreturns?type=r1&start={self.request.params['start']}&end={self.request.params['end']}",
        header,
        self.request,
    )
    wb = gst_r1_template(result["gkdata"])

    output = io.BytesIO()
    wb.save(output)
    contents = output.getvalue()
    output.close()
    headerList = {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Length": len(contents),
        "Content-Disposition": "attachment; filename=report.xlsx",
        "X-Content-Type-Options": "nosniff",
        "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
    }
    # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}
    return Response(contents, headerlist=list(headerList.items()))


def summary(self):
    """Returns GST Summary as spreadsheet

    params
    ======
    calculatefrom = yyyy-mm-dd
    calculateto = yyyy-mm-dd
    statename = String (eg: Delhi)
    orgname = String
    """
    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        gkdata = {
            "startdate": self.request.params["calculatefrom"],
            "enddate": self.request.params["calculateto"],
            "statename": self.request.params["statename"],
        }
        result = gk_api(
            f"/report?type=GSTCalc&startdate={gkdata['startdate']}&enddate={gkdata['enddate']}&statename={gkdata['statename']}",
            header,
            self.request,
        )["gkresult"]

        data = result

        lenSGSTin = len(result["sgstin"])
        lenSGSTout = len(result["sgstout"])
        lenCGSTin = len(result["cgstin"])
        lenCGSTout = len(result["cgstout"])
        lenIGSTin = len(result["igstin"])
        lenIGSTout = len(result["igstout"])
        lenCESSin = len(result["cessin"])
        lenCESSout = len(result["cessout"])
        gstsmwb = openpyxl.Workbook()
        sheet = gstsmwb.active
        sheet.title = "GST Summary "
        ft = Font(name="Liberation Serif", size="12")
        sheet.merge_cells("A1:E2")
        sheet.column_dimensions["A"].width = 8
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A1"] = str(self.request.params["orgname"])
        sheet.merge_cells("A3:E3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"] = "Statement of GST Calculation "
        sheet.merge_cells("A4:E4")
        sheet["A4"].font = Font(name="Liberation Serif", size="12", italic=True)
        sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A4"] = (
            "State :"
            + gkdata["statename"]
            + "                  "
            + "Peroid :"
            + gkdata["startdate"]
            + "to"
            + gkdata["enddate"]
        )

        sheet.merge_cells("A6:A7")
        sheet["A6"].font = Font(name="Liberation Serif", size="12", bold=True)
        sheet["A6"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A6"] = "Type of Tax"

        sheet.merge_cells(start_row=6, start_column=2, end_row=7, end_column=3)
        sheet["B6"].font = Font(name="Liberation Serif", size="12", bold=True)
        sheet["B6"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["B6"] = "Tax Amount"

        sheet.merge_cells("D6:E6")
        sheet["D6"].font = Font(name="Liberation Serif", size="12", bold=True)
        sheet["D6"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["D6"] = "Net Tax Amount"

        sheet["D7"] = "Payable"
        sheet["D7"].font = Font(name="Liberation Serif", size="12", bold=True)
        sheet["D7"].alignment = Alignment(horizontal="center", vertical="center")

        sheet["E7"] = "Carried Forward "
        sheet["E7"].font = Font(name="Liberation Serif", size="12", bold=True)
        sheet["E7"].alignment = Alignment(horizontal="center", vertical="center")

        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 15
        sheet.column_dimensions["E"].width = 15
        # set row=r count and column =c count
        c = 1
        r = 8

        if lenSGSTin > 0:
            t = sheet.cell(row=r, column=c, value="SGST")
            t.font = Font(name="Liberation Serif", bold=True)
            o = sheet.cell(row=r + 1, column=c, value="    Input Tax")
            o.font = ft
            r = r + 2
            # this variables is for indentifying last row of that tax type . We have to write total of each tax type at the last row of mentioned tax type.
            n = 1
            for sgstinacc in data["sgstin"]:
                a = sheet.cell(row=r, column=c, value="        " + sgstinacc)
                a.font = Font(name="Liberation Serif", italic=True)
                v = sheet.cell(
                    row=r,
                    column=c + 1,
                    value=float("%.2f" % float(data["sgstin"][sgstinacc])),
                ).number_format = "0.00"
                if n == lenSGSTin:
                    sheet.cell(
                        row=r,
                        column=c + 2,
                        value=float("%.2f" % float(data["totalSGSTIn"])),
                    ).number_format = "0.00"

                n = n + 1
                r = r + 1

        if lenSGSTout > 0:
            o = sheet.cell(row=r + 1, column=c, value="    Output Tax")
            o.font = ft
            r = r + 2
            n = 1
            for sgstoutacc in data["sgstout"]:
                a = sheet.cell(row=r, column=c, value="        " + sgstoutacc)
                a.font = Font(name="Liberation Serif", italic=True)
                v = sheet.cell(
                    row=r,
                    column=c + 1,
                    value=float("%.2f" % float(data["sgstout"][sgstoutacc])),
                ).number_format = "0.00"
                if n == lenSGSTout:
                    sheet.cell(
                        row=r,
                        column=c + 2,
                        value=float("%.2f" % float(data["totalSGSTOut"])),
                    ).number_format = "0.00"
                    if "sgstpayable" in data:
                        sheet.cell(
                            row=r,
                            column=c + 3,
                            value=float("%.2f" % float(data["sgstpayable"])),
                        ).number_format = "0.00"
                    if "sgstcrdfwd" in data:
                        sheet.cell(
                            row=r,
                            column=c + 4,
                            value=float("%.2f" % float(data["sgstcrdfwd"])),
                        ).number_format = "0.00"
                n = n + 1
                r = r + 1

        if lenCGSTin > 0:
            t = sheet.cell(row=r, column=c, value="CGST")
            t.font = Font(name="Liberation Serif", bold=True)
            i = sheet.cell(row=r + 1, column=c, value="    Input Tax")
            i.font = Font(name="Liberation Serif")
            r = r + 2
            n = 1
            for cgstinacc in data["cgstin"]:
                a = sheet.cell(row=r, column=c, value="        " + cgstinacc)
                a.font = Font(name="Liberation Serif", italic=True)
                v = sheet.cell(
                    row=r,
                    column=c + 1,
                    value=float("%.2f" % float(data["cgstin"][cgstinacc])),
                ).number_format = "0.00"
                if n == lenCGSTin:
                    sheet.cell(
                        row=r,
                        column=c + 2,
                        value=float("%.2f" % float(data["totalCGSTIn"])),
                    ).number_format = "0.00"

                n = n + 1
                r = r + 1

        if lenCGSTout > 0:
            o = sheet.cell(row=r + 1, column=c, value="    Output Tax")
            o.font = ft
            r = r + 2
            n = 1
            for cgstoutacc in data["cgstout"]:
                a = sheet.cell(row=r, column=c, value="        " + cgstoutacc)
                a.font = Font(name="Liberation Serif", italic=True)
                v = sheet.cell(
                    row=r,
                    column=c + 1,
                    value=float("%.2f" % float(data["cgstout"][cgstoutacc])),
                ).number_format = "0.00"
                if n == lenCGSTout:
                    sheet.cell(
                        row=r,
                        column=c + 2,
                        value=float("%.2f" % float(data["totalCGSTOut"])),
                    ).number_format = "0.00"
                    if "cgstpayable" in data:
                        sheet.cell(
                            row=r,
                            column=c + 3,
                            value=float("%.2f" % float(data["cgstpayable"])),
                        ).number_format = "0.00"
                    if "cgstcrdfwd" in data:
                        sheet.cell(
                            row=r,
                            column=c + 4,
                            value=float("%.2f" % float(data["cgstcrdfwd"])),
                        ).number_format = "0.00"

                n = n + 1
                r = r + 1

        if lenIGSTin > 0:
            t = sheet.cell(row=r, column=c, value="IGST")
            t.font = Font(name="Liberation Serif", bold=True)
            i = sheet.cell(row=r + 1, column=c, value="    Input Tax")
            i.font = ft
            r = r + 2
            n = 1
            for igstinacc in data["igstin"]:
                a = sheet.cell(row=r, column=c, value="        " + igstinacc)
                a.font = Font(name="Liberation Serif", italic=True)
                v = sheet.cell(
                    row=r,
                    column=c + 1,
                    value=float("%.2f" % float(data["igstin"][igstinacc])),
                ).number_format = "0.00"
                if n == lenIGSTin:
                    sheet.cell(
                        row=r,
                        column=c + 2,
                        value=float("%.2f" % float(data["totalIGSTIn"])),
                    ).number_format = "0.00"

                n = n + 1
                r = r + 1

        if lenIGSTout > 0:
            o = sheet.cell(row=r + 1, column=c, value="    Output Tax")
            o.font = ft
            r = r + 2
            n = 1
            for igstoutacc in data["igstout"]:
                a = sheet.cell(row=r, column=c, value="        " + igstoutacc)
                a.font = Font(name="Liberation Serif", italic=True)
                v = sheet.cell(
                    row=r,
                    column=c + 1,
                    value=float("%.2f" % float(data["igstout"][igstoutacc])),
                ).number_format = "0.00"
                if n == lenIGSTout:
                    sheet.cell(
                        row=r,
                        column=c + 2,
                        value=float("%.2f" % float(data["totalIGSTOut"])),
                    ).number_format = "0.00"
                    if "IgstPayable" in data:
                        sheet.cell(
                            row=r,
                            column=c + 3,
                            value=float("%.2f" % float(data["IgstPayable"])),
                        ).number_format = "0.00"
                    if "IgstCrdFwd" in data:
                        sheet.cell(
                            row=r,
                            column=c + 4,
                            value=float("%.2f" % float(data["IgstCrdFwd"])),
                        ).number_format = "0.00"

                n = n + 1
                r = r + 1

        if lenCESSin > 0:
            t = sheet.cell(row=r, column=c, value="CESS")
            t.font = Font(name="Liberation Serif", bold=True)
            i = sheet.cell(row=r + 1, column=c, value="    Input Tax")
            i.font = ft
            r = r + 2
            n = 1
            for cessacc in data["cessin"]:
                a = sheet.cell(row=r, column=c, value="        " + cessacc)
                a.font = Font(name="Liberation Serif", italic=True)
                v = sheet.cell(
                    row=r,
                    column=c + 1,
                    value=float("%.2f" % float(data["cessin"][cessacc])),
                ).number_format = "0.00"
                if n == lenCESSin:
                    sheet.cell(
                        row=r,
                        column=c + 2,
                        value=float("%.2f" % float(data["totalCESSIn"])),
                    ).number_format = "0.00"

                n = n + 1
                r = r + 1

        if lenCESSout > 0:
            o = sheet.cell(row=r + 1, column=c, value="    Output Tax")
            o.font = ft
            r = r + 2
            n = 1
            for cessoutacc in data["cessout"]:
                a = sheet.cell(row=r, column=c, value="        " + cessoutacc)
                a.font = Font(name="Liberation Serif", italic=True)
                v = sheet.cell(
                    row=r,
                    column=c + 1,
                    value=float("%.2f" % float(data["cessout"][cessoutacc])),
                ).number_format = "0.00"
                if n == lenCESSout:
                    sheet.cell(
                        row=r,
                        column=c + 2,
                        value=float("%.2f" % float(data["totalCESSOut"])),
                    ).number_format = "0.00"
                    if "cesspayable" in data:
                        sheet.cell(
                            row=r,
                            column=c + 3,
                            value=float("%.2f" % float(data["cesspayable"])),
                        ).number_format = "0.00"
                    if "cessCrdFwd" in data:
                        sheet.cell(
                            row=r,
                            column=c + 4,
                            value=float("%.2f" % float(data["cessCrdFwd"])),
                        ).number_format = "0.00"

                n = n + 1
                r = r + 1

        output = io.BytesIO()
        gstsmwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/ '}
        return Response(contents, headerlist=list(headerList.items()))

    except Exception as e:
        print(e)
        return {"gkstatus": 3}
