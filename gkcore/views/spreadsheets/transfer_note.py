"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Survesh" <123survesh@gmail.com>
"Sai Karthik"<kskarthik@disroot.org>

"""
from gkcore.models.meta import gk_api
from pyramid.response import Response

# Spreadsheet libraries
import openpyxl
from openpyxl.styles import Font, Alignment

# from io import BytesIO
import io


def all_transfer_notes(self):
    """List of all transfer notes

    This function returns a spreadsheet form of List of Transfer Notes Report.
    The spreadsheet in XLSX format is generated by the frontend.

    params
    ======
    fystart = dd-mm-yyyy
    fyend = dd-mm-yyyy
    orgname = int
    startdate = dd-mm-yyyy
    enddate = dd-mm-yyyy
    goid(optional) = int
    """

    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        orgname = str(self.request.params["orgname"])
        orgname += " (FY: " + fystart + " to " + fyend + ")"
        startDate = str(self.request.params["startdate"])
        endDate = str(self.request.params["enddate"])
        godownname = ""
        godownaddress = ""
        goid = 0
        transfernotewb = openpyxl.Workbook()
        sheet = transfernotewb.active
        sheet.title = "List of Transfer Notes"
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 24
        sheet.column_dimensions["F"].width = 20
        sheet.column_dimensions["G"].width = 16
        sheet.column_dimensions["H"].width = 14
        sheet.merge_cells("A1:H2")
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet["A1"] = orgname
        sheet.merge_cells("A3:H3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"] = "List of Transfer Notes"
        sheet.merge_cells("A4:H4")
        sheet["A4"] = "Period: " + startDate + " to " + endDate
        sheet["A4"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
        titlerow = 5
        # If an id of a godown is received it will give all transfernotes involving that godown with godownname and godownaddress.
        if "goid" in self.request.params:
            goid = int(self.request.params["goid"])
            transfernotes = gk_api(
                f"/transfernote?type=list&startdate={startDate}&enddate={endDate}&goid={goid}",
                header,
                self.request,
            )
            godown = gk_api(
                f"/godown?qty=single&goid={int(self.request.params['goid'])}",
                header,
                self.request,
            )
            godownname = godown["gkresult"]["goname"]
            godownaddress = godown["gkresult"]["goaddr"]
            nameofgodown = (
                "Name of Godown: "
                + godownname
                + "           Godown Address: "
                + godownaddress
            )
            sheet.merge_cells("A5:H5")
            sheet["A5"] = nameofgodown
            sheet["A5"].font = Font(name="Liberation Serif", size="14", bold=True)
            sheet["A5"].alignment = Alignment(horizontal="center", vertical="center")
            titlerow = 6
        else:
            transfernotes = gk_api(
                f"/transfernote?type=list&startdate={startDate}&enddate={endDate}",
                header,
                self.request,
            )
        transfernotes = transfernotes["gkresult"]
        sheet["A" + str(titlerow)] = "Sr. No."
        sheet["B" + str(titlerow)] = "TN No."
        sheet["C" + str(titlerow)] = "Date"
        sheet["D" + str(titlerow)] = "Dispatch From"
        sheet["E" + str(titlerow)] = "To be Delivered At"
        sheet["F" + str(titlerow)] = "Products"
        sheet["G" + str(titlerow)] = "Quantity"
        sheet["H" + str(titlerow)] = "Status"
        sheet["A" + str(titlerow)].alignment = Alignment(horizontal="center")
        sheet["B" + str(titlerow)].alignment = Alignment(horizontal="center")
        sheet["C" + str(titlerow)].alignment = Alignment(horizontal="center")
        sheet["D" + str(titlerow)].alignment = Alignment(horizontal="center")
        sheet["E" + str(titlerow)].alignment = Alignment(horizontal="center")
        sheet["F" + str(titlerow)].alignment = Alignment(horizontal="center")
        sheet["G" + str(titlerow)].alignment = Alignment(horizontal="right")
        sheet["H" + str(titlerow)].alignment = Alignment(horizontal="center")
        sheet["A" + str(titlerow)].font = Font(
            name="Liberation Serif", size=12, bold=True
        )
        sheet["B" + str(titlerow)].font = Font(
            name="Liberation Serif", size=12, bold=True
        )
        sheet["C" + str(titlerow)].font = Font(
            name="Liberation Serif", size=12, bold=True
        )
        sheet["D" + str(titlerow)].font = Font(
            name="Liberation Serif", size=12, bold=True
        )
        sheet["E" + str(titlerow)].font = Font(
            name="Liberation Serif", size=12, bold=True
        )
        sheet["F" + str(titlerow)].font = Font(
            name="Liberation Serif", size=12, bold=True
        )
        sheet["G" + str(titlerow)].font = Font(
            name="Liberation Serif", size=12, bold=True
        )
        sheet["H" + str(titlerow)].font = Font(
            name="Liberation Serif", size=12, bold=True
        )
        row = titlerow + 1
        # Looping each dictionaries in list transfernotes to store data in cells and apply styles.
        for transfernote in transfernotes:
            sheet["A" + str(row)] = transfernote["srno"]
            sheet["A" + str(row)].alignment = Alignment(horizontal="center")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["B" + str(row)] = transfernote["transfernoteno"]
            sheet["B" + str(row)].alignment = Alignment(horizontal="center")
            sheet["B" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["C" + str(row)] = transfernote["transfernotedate"]
            sheet["C" + str(row)].alignment = Alignment(horizontal="center")
            sheet["C" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["D" + str(row)] = transfernote["fromgodown"]
            sheet["D" + str(row)].alignment = Alignment(horizontal="left")
            sheet["D" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["E" + str(row)] = transfernote["togodown"]
            sheet["E" + str(row)].alignment = Alignment(horizontal="left")
            sheet["E" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            subrow = row
            for productqty in transfernote["productqty"]:
                sheet["F" + str(subrow)] = productqty["productdesc"]
                sheet["F" + str(subrow)].alignment = Alignment(horizontal="left")
                sheet["F" + str(subrow)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                sheet["G" + str(subrow)] = (
                    productqty["quantity"] + " " + productqty["uom"]
                )
                sheet["G" + str(subrow)].alignment = Alignment(horizontal="right")
                sheet["G" + str(subrow)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                subrow += 1
            if transfernote["receivedflag"]:
                sheet["H" + str(row)] = "Received"
                sheet["H" + str(row)].alignment = Alignment(horizontal="center")
                sheet["H" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            else:
                sheet["H" + str(row)] = "Pending"
                sheet["H" + str(row)].alignment = Alignment(horizontal="center")
                sheet["H" + str(row)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
            if subrow == row:
                row += 1
            else:
                row = subrow

        output = io.BytesIO()
        transfernotewb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print(e)
        return {"gkstatus": 3}
