"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020,2019 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Bhavesh Bawadhane" <bbhavesh07@gmail.com>
'Prajkta Patkar' <prajkta@gnukhata.in>
'Sai Karthik' <kskarthik@disroot.org>
"Survesh" <123survesh@gmail.com>
"""

# from pyramid.view import view_config
import json

# from datetime import datetime
# from pyramid.renderers import render_to_response
from pyramid.response import Response
from pyramid.request import Request
import io

# from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles import DEFAULT_FONT

# import calendar


def print_profit_loss(self):
    """
    Generates profit & loss statement for given date range

    params
    ======

    from
    to
    fystart
    orgname
    orgtype
    """
    calculatefrom = self.request.params["from"]
    calculateto = self.request.params["to"]
    orgtype = self.request.params["orgtype"]
    header = {"gktoken": self.request.headers["gktoken"]}
    fystart = str(self.request.params["fystart"])
    fyend = str(self.request.params["fyend"])
    orgname = str(self.request.params["orgname"])
    req = Request.blank(
        "/reports/profit-loss?calculatefrom=%s&calculateto=%s"
        % (calculatefrom, calculateto),
        headers=header,
    )
    result = json.loads(self.request.invoke_subrequest(req).text)["gkresult"]

    # A workbook is opened.
    wb = openpyxl.Workbook()
    # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
    sheet = wb.active
    DEFAULT_FONT.name = "Liberation Serif"
    sheet.merge_cells("A1:H2")
    sheet["A1"].font = Font(size="16", bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
    sheet.merge_cells("A3:H3")
    # Title of the sheet and width of columns are set.
    if orgtype == "Profit Making":
        title = "Profit and Loss"
    else:
        title = "Income and Expenditure"

    sheet.title = title
    sheet["A3"] = f"{title} from {fystart} to {fyend}"
    sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A3"].font = Font(size="13", bold=True)
    sheet.column_dimensions["A"].width = 50
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 50
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 16

    trading_left = format_list(result["trading_left"])
    trading_right = format_list(result["trading_right"])
    pnl_left = format_list(result["pnl_left"])
    pnl_right = format_list(result["pnl_right"])

    trading_start_row = 5

    sheet[f"A{trading_start_row}"] = "Particulars"
    sheet[f"A{trading_start_row}"].alignment = Alignment(horizontal="left")
    sheet[f"A{trading_start_row}"].font = Font(size="12", bold=True)
    sheet[f"D{trading_start_row}"] = "Particulars"
    sheet[f"D{trading_start_row}"].alignment = Alignment(horizontal="left")
    sheet[f"D{trading_start_row}"].font = Font(size="12", bold=True)

    trading_left_last_row, trading_left_total = update_sheet(
        sheet, trading_left, trading_start_row+1, "left"
    )
    trading_right_last_row, trading_right_total = update_sheet(
        sheet, trading_right, trading_start_row+1, "right"
    )
    trading_last_row = max(trading_left_last_row, trading_right_last_row)

    sheet[f"A{trading_last_row}"] = sheet[f"D{trading_last_row}"] = "TOTAL"
    sheet[f"C{trading_last_row}"] = trading_left_total
    sheet[f"C{trading_last_row}"] = trading_left_total
    sheet[f"F{trading_last_row}"] = trading_right_total
    for cell in [sheet[f"A{trading_last_row}"], sheet[f"D{trading_last_row}"]]:
        cell.font = Font(size="13", bold=True)
    for cell in [sheet[f"C{trading_last_row}"], sheet[f"F{trading_last_row}"]]:
        cell.font = Font(size="13", bold=True, u="doubleAccounting")

    pnl_start_row = trading_last_row + 2

    sheet[f"A{pnl_start_row}"] = "Particulars"
    sheet[f"A{pnl_start_row}"].alignment = Alignment(horizontal="left")
    sheet[f"A{pnl_start_row}"].font = Font(size="12", bold=True)
    sheet[f"D{pnl_start_row}"] = "Particulars"
    sheet[f"D{pnl_start_row}"].alignment = Alignment(horizontal="left")
    sheet[f"D{pnl_start_row}"].font = Font(size="12", bold=True)

    pnl_left_last_row, pnl_left_total = update_sheet(
        sheet, pnl_left, pnl_start_row+1, "left"
    )
    pnl_right_last_row, pnl_right_total = update_sheet(
        sheet, pnl_right, pnl_start_row+1, "right"
    )
    pnl_last_row = max(pnl_left_last_row, pnl_right_last_row)

    sheet[f"A{pnl_last_row}"] = sheet[f"D{pnl_last_row}"] = "TOTAL"
    sheet[f"C{pnl_last_row}"] = pnl_left_total
    sheet[f"F{pnl_last_row}"] = pnl_right_total
    for cell in [sheet[f"A{pnl_last_row}"], sheet[f"D{pnl_last_row}"]]:
        cell.font = Font(size="13", bold=True)
    for cell in [sheet[f"C{pnl_last_row}"], sheet[f"F{pnl_last_row}"]]:
        cell.font = Font(size="13", bold=True, u="doubleAccounting")

    for row in range(1, pnl_last_row+2):
        sheet["B{}".format(row)].number_format = '#,##0.00'
        sheet["C{}".format(row)].number_format = '#,##0.00'
        sheet["E{}".format(row)].number_format = '#,##0.00'
        sheet["F{}".format(row)].number_format = '#,##0.00'

    output = io.BytesIO()
    wb.save(output)
    contents = output.getvalue()
    output.close()
    headerList = {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Length": len(contents),
        "Content-Disposition": "attachment; filename=report.xlsx",
        "X-Content-Type-Options": "nosniff",
        "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
    }

    return Response(contents, headerlist=list(headerList.items()))


def update_sheet(sheet, report_rows, start_row, report_side):
    if report_side == "left":
        label_column = "A"
        sub_value_column = "B"
        value_column = "C"
    else:
        label_column = "D"
        sub_value_column = "E"
        value_column = "F"
    row_no = start_row
    total = 0
    for row in report_rows:
        if row["type"] == "total":
            total = row["amount"]
            continue
        row_no += 1
        sheet[f"{label_column}{row_no}"] = row["name"]
        sheet[f"{value_column}{row_no}"] = row["amount"]
        sheet[f"{value_column}{row_no}"].font = Font(
            name="Liberation Serif", size="12", bold=True
        )

        if row.get("accounts"):
            for account in row["accounts"]:
                row_no += 1
                sheet[f"{label_column}{row_no}"] = "    " + account["accountname"]
                sheet[f"{sub_value_column}{row_no}"] = account["amount"]
    return row_no, total


def format_list(rows):
    formated_list = []
    for row in rows:
        if row["type"] == "account" and row["subgroupcode"]:
            continue
        formated_list.append(row)
    return formated_list
