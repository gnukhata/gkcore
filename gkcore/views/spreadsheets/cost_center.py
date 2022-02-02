"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020, 2021 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors
============
Survesh VRL <123survesh@gmail.com>
Sai Karthik <kskarthik@disroot.org>

"""


from gkcore.models.meta import gk_api

from pyramid.response import Response

# Spreadsheet libraries
import io
import openpyxl
from openpyxl.styles import Font, Alignment


def statement(self):
    """
    This function returns a spreadsheet form of Cost Center Statement Report.
    The spreadsheet in XLSX format"

    params
    ======

    fystart =
    fyend =
    orgname =
    projectname =
    projectcode =
    calculatefrom =
    calculateto =
    """
    try:
        calculateto = self.request.params["calculateto"]
        financialstart = self.request.params["fystart"]
        projectcode = int(self.request.params["projectcode"])
        projectname = self.request.params["projectname"]
        header = {"gktoken": self.request.headers["gktoken"]}
        result = gk_api(
            "/report?type=projectstatement&calculateto=%s&financialstart=%s&projectcode=%d"
            % (calculateto, financialstart, projectcode),
            header,
            self.request,
        )["gkresult"]
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        fystart = fystart[8:10] + fystart[4:8] + fystart[0:4]
        fyend = fyend[8:10] + fyend[4:8] + fyend[0:4]
        calculateto = str(self.request.params["calculateto"])
        calculateto = calculateto[8:10] + calculateto[4:8] + calculateto[0:4]
        orgname = str(self.request.params["orgname"])
        projstmtwb = openpyxl.Workbook()
        sheet = projstmtwb.active
        sheet.title = "Project Statement (" + projectname + ")"
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 16
        sheet.column_dimensions["E"].width = 16
        sheet.merge_cells("A1:E2")
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet.merge_cells("A3:E3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"] = (
            "Statement for :"
            + projectname
            + " ("
            + fystart
            + " to "
            + calculateto
            + ")"
        )
        sheet["A4"] = "Sr. No. "
        sheet["B4"] = "Account Name"
        sheet["C4"] = "Group Name"
        sheet["D4"] = "Sub-Group Name"
        sheet["E4"] = "Outgoing"
        sheet["F4"] = "Incoming"
        sheet["A4"].alignment = Alignment(horizontal="center")
        sheet["B4"].alignment = Alignment(horizontal="left")
        sheet["C4"].alignment = Alignment(horizontal="left")
        sheet["D4"].alignment = Alignment(horizontal="left")
        sheet["E4"].alignment = Alignment(horizontal="right")
        sheet["F4"].alignment = Alignment(horizontal="right")
        sheet["A4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["B4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["C4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["D4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["E4"].font = Font(name="Liberation Serif", size=12, bold=True)
        sheet["F4"].font = Font(name="Liberation Serif", size=12, bold=True)
        row = 5
        # Looping each dictionaries in list result to store data in cells and apply styles.
        for transaction in result:
            sheet["A" + str(row)] = transaction["srno"]
            sheet["A" + str(row)].alignment = Alignment(horizontal="center")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["B" + str(row)] = transaction["accountname"]
            sheet["B" + str(row)].alignment = Alignment(horizontal="left")
            sheet["B" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["C" + str(row)] = transaction["groupname"]
            sheet["C" + str(row)].alignment = Alignment(horizontal="left")
            sheet["C" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["D" + str(row)] = transaction["subgroupname"]
            sheet["D" + str(row)].alignment = Alignment(horizontal="left")
            sheet["D" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["E" + str(row)] = float("%.2f" % float(transaction["totalout"]))
            sheet["E" + str(row)].number_format = "0.00"
            sheet["E" + str(row)].alignment = Alignment(horizontal="right")
            sheet["E" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["F" + str(row)] = float("%.2f" % float(transaction["totalin"]))
            sheet["F" + str(row)].number_format = "0.00"
            sheet["F" + str(row)].alignment = Alignment(horizontal="right")
            sheet["F" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            row += 1
        output = io.BytesIO()
        projstmtwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/ '}
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print(e)
        return {"gkstatus": 3}
