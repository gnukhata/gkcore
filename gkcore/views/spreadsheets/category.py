"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020, 2021 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors
============
Survesh VRL <123survesh@gmail.com>
Sai Karthik <kskarthik@disroot.org>

"""
from gkcore.models.meta import gk_api
from gkcore.models.meta import gk_api
from pyramid.response import Response

# Spreadsheet libraries
import io
import openpyxl
from openpyxl.styles import Font, Alignment


def all_categories(self):
    try:
        header = {"gktoken": self.request.headers["gktoken"]}
        result = gk_api("/categories", header, self.request)["gkresult"]
        fystart = str(self.request.params["fystart"])
        fyend = str(self.request.params["fyend"])
        orgname = str(self.request.params["orgname"])
        # A workbook is opened.
        categorywb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = categorywb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "List of Categories"
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 24
        sheet.column_dimensions["C"].width = 24
        sheet.column_dimensions["D"].width = 14
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells("A1:G2")
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        # Organisation name and financial year are displayed.
        sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
        sheet.merge_cells("A3:G3")
        sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"] = "List of Categories"
        sheet.merge_cells("A3:G3")
        sheet["A4"] = "Sr. No."
        sheet["B4"] = "Category"
        sheet["C4"] = "Sub-Category"
        sheet["D4"] = "Status"
        titlerow = sheet.row_dimensions[4]
        titlerow.font = Font(name="Liberation Serif", size=12, bold=True)
        row = 5
        srno = 1
        for category in result:
            sheet["A" + str(row)] = srno
            sheet["A" + str(row)].alignment = Alignment(horizontal="left")
            sheet["A" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            sheet["B" + str(row)] = category["categoryname"]
            sheet["B" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            children = gk_api(
                "/categories?type=children&categorycode=%d"
                % (int(category["categorycode"])),
                header,
                self.request,
            )["gkresult"]
            subrow = row
            for child in children:
                sheet["C" + str(subrow)] = child["categoryname"]
                sheet["C" + str(subrow)].font = Font(
                    name="Liberation Serif", size="12", bold=False
                )
                subrow += 1
            sheet["D" + str(row)] = category["categorystatus"]
            sheet["D" + str(row)].font = Font(
                name="Liberation Serif", size="12", bold=False
            )
            if subrow == row:
                row += 1
            else:
                row = subrow
            srno += 1
        output = io.BytesIO()
        categorywb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Length": len(contents),
            "Content-Disposition": "attachment; filename=report.xlsx",
            "X-Content-Type-Options": "nosniff",
            "Set-Cookie": "fileDownload=true; path=/ [;HttpOnly]",
        }
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except Exception as e:
        print(e)
        return {"gkstatus": 3}
