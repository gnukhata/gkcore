"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Survesh" <123survesh@gmail.com>
"Sai Karthik"<kskarthik@disroot.org>

"""

from gkcore import eng, enumdict
from gkcore.views.api_login import authCheck
from pyramid.response import Response
from gkcore.views.api_reports import getBalanceSheet

# Spreadsheet libraries
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import Color

RED = Color(rgb="FF0000")


def print_bal_sheet(self):

    """
    Returns the spreadsheet for Conventional Balancesheet
    In the Conventional format, capital and liabilities are shown on left side and assets on the right side.

    Params
        type            = conv_bal_sheet
        calculateto     = Calculate balance sheet to date (yyyy-mm-dd)
        calculatefrom   = Calculate balance sheet from date (yyyy-mm-dd)
        baltype         = balance sheet type
        orgname         = Organisation name
        fystart         = Financial start date (dd-mm-yyyy)
        fyend           = Financial end date (dd-mm-yyyy)
        orgtype         = Organisation type, profitable or not for profit
    """

    try:
        token = self.request.headers["gktoken"]
    except:
        return {"gkstatus": enumdict["UnauthorisedAccess"]}
    authDetails = authCheck(token)
    if authDetails["auth"] == False:
        return {"gkstatus": enumdict["UnauthorisedAccess"]}
    else:
        try:
            self.con = eng.connect()
            calculateto = self.request.params["calculateto"]
            calculatefrom = self.request.params["calculatefrom"]
            balType = int(self.request.params["baltype"])
            fystart = str(self.request.params["fystart"])
            orgname = str(self.request.params["orgname"])
            fyend = str(self.request.params["fyend"])
            orgtype = str(self.request.params["orgtype"])
            result = getBalanceSheet(
                self.con,
                authDetails["orgcode"],
                calculateto,
                calculatefrom,
                balType,
            )
            calculateto = calculateto[8:10] + calculateto[4:8] + calculateto[0:4]
            sources = result["leftlist"]
            applications = result["rightlist"]
            conventionalwb = openpyxl.Workbook()
            sheet = conventionalwb.active
            sheet.title = "Conventional Balance Sheet"
            sheet.column_dimensions["A"].width = 30
            sheet.column_dimensions["B"].width = 12
            sheet.column_dimensions["C"].width = 12
            sheet.column_dimensions["D"].width = 12
            sheet.column_dimensions["E"].width = 30
            sheet.column_dimensions["F"].width = 12
            sheet.column_dimensions["G"].width = 12
            sheet.column_dimensions["H"].width = 12
            sheet.merge_cells("A1:H2")
            sheet["A1"].font = Font(name="Liberation Serif", size="16", bold=True)
            sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
            sheet["A1"] = orgname + " (FY: " + fystart + " to " + fyend + ")"
            sheet.merge_cells("A3:H3")
            if orgtype == "Profit Making":
                sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
                sheet["A3"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                sheet["A3"] = (
                    "Conventional Balance Sheet from "
                    + calculatefrom
                    + " to "
                    + calculateto
                )
                sheet["A4"] = "Capital and Liabilities"
                sheet["A4"].font = Font(name="Liberation Serif", size="12", bold=True)
            if orgtype == "Not For Profit":
                sheet["A3"].font = Font(name="Liberation Serif", size="14", bold=True)
                sheet["A3"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                sheet["A3"] = "Conventional Statement of Affairs as on " + calculateto
                sheet["A4"] = "Corpus and Liabilities"
                sheet["A4"].font = Font(name="Liberation Serif", size="12", bold=True)
            sheet["D4"] = "Amount"
            sheet["D4"].alignment = Alignment(horizontal="right")
            sheet["D4"].font = Font(name="Liberation Serif", size="12", bold=True)
            sheet["E4"] = "Property and Assets"
            sheet["E4"].alignment = Alignment(horizontal="left")
            sheet["E4"].font = Font(name="Liberation Serif", size="12", bold=True)
            sheet["H4"] = "Amount"
            sheet["H4"].alignment = Alignment(horizontal="right")
            sheet["H4"].font = Font(name="Liberation Serif", size="12", bold=True)
            """
            """
            """
            Looping each dictionaries in list sources to store data in cells and apply styles.
            If the advflag = 1 then 'amount' will be displayed in 'RED' color
            """
            row = 4
            for record in sources:
                if record["groupAccname"] != "":
                    if record["groupAccname"] != "Sources:":
                        if (
                            record["groupAccname"] == "Total"
                            or record["groupAccname"] == "Sources:"
                            or record["groupAccname"] == "Difference"
                        ):
                            sheet["A" + str(row)] = record["groupAccname"].upper()
                            sheet["A" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=True
                            )
                        elif (
                            record["groupAccflag"] == "" and record["subgroupof"] != ""
                        ):
                            sheet["A" + str(row)] = (
                                "            " + record["groupAccname"]
                            )
                            sheet["A" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                        elif record["groupAccflag"] == 1:
                            sheet["A" + str(row)] = (
                                "                        " + record["groupAccname"]
                            )
                            sheet["A" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                        elif record["groupAccflag"] == 2:
                            sheet["A" + str(row)] = (
                                "                        " + record["groupAccname"]
                            )
                            sheet["A" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                        else:
                            sheet["A" + str(row)] = record["groupAccname"].upper()
                            sheet["A" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                    if record["groupAccflag"] == 2 or record["groupAccflag"] == 1:
                        if record["advflag"] == 1:
                            sheet["B" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["B" + str(row)].number_format = "0.00"
                            sheet["B" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["B" + str(row)].font = Font(
                                name="Liberation Serif",
                                size="12",
                                bold=True,
                                color=RED,
                            )
                        else:
                            sheet["B" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["B" + str(row)].number_format = "0.00"
                            sheet["B" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["B" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                    elif record["groupAccflag"] == "" and record["subgroupof"] != "":
                        if record["advflag"] == 1:
                            sheet["C" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["C" + str(row)].number_format = "0.00"
                            sheet["C" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["C" + str(row)].font = Font(
                                name="Liberation Serif",
                                size="12",
                                bold=True,
                                color=RED,
                            )
                        else:
                            sheet["C" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["C" + str(row)].number_format = "0.00"
                            sheet["C" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["C" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                    else:
                        if record["advflag"] == 1:
                            sheet["D" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["D" + str(row)].number_format = "0.00"
                            sheet["D" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["D" + str(row)].font = Font(
                                name="Liberation Serif",
                                size="12",
                                bold=True,
                                color=RED,
                            )
                        else:
                            if record["amount"] != "":
                                sheet["D" + str(row)] = float(
                                    "%.2f" % float(record["amount"])
                                )
                                sheet["D" + str(row)].number_format = "0.00"
                                sheet["D" + str(row)].alignment = Alignment(
                                    horizontal="right"
                                )
                                sheet["D" + str(row)].font = Font(
                                    name="Liberation Serif", size="12", bold=True
                                )
                    row += 1
            """
            """
            """
            Looping each dictionaries in list applications to store data in cells and apply styles.
            If the advflag = 1 then 'amount' will be displayed in 'RED' color
            """
            row = 4
            for record in applications:
                if record["groupAccname"] != "":
                    if record["groupAccname"] != "Applications:":
                        if (
                            record["groupAccname"] == "Total"
                            or record["groupAccname"] == "Sources:"
                            or record["groupAccname"] == "Difference"
                        ):
                            sheet["E" + str(row)] = record["groupAccname"].upper()
                            sheet["E" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=True
                            )
                        elif (
                            record["groupAccflag"] == "" and record["subgroupof"] != ""
                        ):
                            sheet["E" + str(row)] = (
                                "            " + record["groupAccname"]
                            )
                            sheet["E" + str(row)].font = Font(
                                name="Liberation Serif", size="12"
                            )
                        elif record["groupAccflag"] == 1:
                            sheet["E" + str(row)] = (
                                "                        " + record["groupAccname"]
                            )
                            sheet["E" + str(row)].font = Font(
                                name="Liberation Serif", size="12"
                            )
                        elif record["groupAccflag"] == 2:
                            sheet["E" + str(row)] = (
                                "                        " + record["groupAccname"]
                            )
                            sheet["E" + str(row)].font = Font(
                                name="Liberation Serif", size="12"
                            )
                        else:
                            sheet["E" + str(row)] = record["groupAccname"].upper()
                            sheet["E" + str(row)].font = Font(
                                name="Liberation Serif", size="12"
                            )

                    if record["groupAccflag"] == 2 or record["groupAccflag"] == 1:
                        if record["advflag"] == 1:
                            sheet["F" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["F" + str(row)].number_format = "0.00"
                            sheet["F" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["F" + str(row)].font = Font(
                                name="Liberation Serif",
                                size="12",
                                bold=True,
                                color=RED,
                            )
                        else:
                            sheet["F" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["F" + str(row)].number_format = "0.00"
                            sheet["F" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["F" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                    elif record["groupAccflag"] == "" and record["subgroupof"] != "":
                        if record["advflag"] == 1:
                            sheet["G" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["G" + str(row)].number_format = "0.00"
                            sheet["G" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["G" + str(row)].font = Font(
                                name="Liberation Serif",
                                size="12",
                                bold=True,
                                color=RED,
                            )
                        else:
                            sheet["G" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["G" + str(row)].number_format = "0.00"
                            sheet["G" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["G" + str(row)].font = Font(
                                name="Liberation Serif", size="12", bold=False
                            )
                    else:
                        if record["advflag"] == 1:
                            sheet["H" + str(row)] = float(
                                "%.2f" % float(record["amount"])
                            )
                            sheet["H" + str(row)].number_format = "0.00"
                            sheet["H" + str(row)].alignment = Alignment(
                                horizontal="right"
                            )
                            sheet["H" + str(row)].font = Font(
                                name="Liberation Serif",
                                size="12",
                                bold=True,
                                color=RED,
                            )
                        else:
                            if record["amount"] != "":
                                sheet["H" + str(row)] = float(
                                    "%.2f" % float(record["amount"])
                                )
                                sheet["H" + str(row)].number_format = "0.00"
                                sheet["H" + str(row)].alignment = Alignment(
                                    horizontal="right"
                                )
                                sheet["H" + str(row)].font = Font(
                                    name="Liberation Serif", size="12", bold=True
                                )
                    row += 1
            output = io.BytesIO()
            conventionalwb.save(output)
            contents = output.getvalue()
            output.close()
            headerList = {
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Length": len(contents),
                "Content-Disposition": "attachment; filename=report.xlsx",
                "X-Content-Type-Options": "nosniff",
                "Set-Cookie": "fileDownload=true ;path=/ [;HttpOnly]",
            }
            # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
            self.con.close()
            return Response(contents, headerlist=list(headerList.items()))
        except:
            #   print(e)
            self.con.close()
            return {"gkstatus": enumdict["ConnectionFailed"]}
